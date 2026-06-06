from io import BytesIO
import re

from openpyxl import load_workbook


OSM_EVENT_ATTENDANCE_MAP = {
    "yes": "Attending",
    "y": "Attending",
    "attending": "Attending",
    "confirmed": "Attending",
    "no": "Not attending",
    "n": "Not attending",
    "not attending": "Not attending",
    "invited": "Invited",
    "": "No response",
}


def clean_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def normalise_label(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def normalise_person_match_value(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def combine_names(first_name, last_name):
    return " ".join(part for part in [clean_cell(first_name), clean_cell(last_name)] if part)


def map_attendance(value):
    key = normalise_label(value)
    return OSM_EVENT_ATTENDANCE_MAP.get(key, clean_cell(value) or "No response")


def parse_osm_event_attendance_export(file_bytes):
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    rows = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        if worksheet.max_row < 2:
            continue

        header_map = {}

        for col in range(1, worksheet.max_column + 1):
            header = normalise_label(worksheet.cell(row=1, column=col).value)
            if header:
                header_map[header] = col

        first_name_col = header_map.get("first name")
        last_name_col = header_map.get("last name") or header_map.get("surname")
        attending_col = header_map.get("attending") or header_map.get("attendance")

        if not first_name_col or not last_name_col or not attending_col:
            continue

        for row_number in range(2, worksheet.max_row + 1):
            first_name = clean_cell(worksheet.cell(row=row_number, column=first_name_col).value)
            last_name = clean_cell(worksheet.cell(row=row_number, column=last_name_col).value)
            attending_raw = clean_cell(worksheet.cell(row=row_number, column=attending_col).value)

            if not first_name and not last_name:
                continue

            rows.append(
                {
                    "sheet_name": sheet_name,
                    "excel_row": row_number,
                    "first_name": first_name,
                    "last_name": last_name,
                    "display_name": combine_names(first_name, last_name),
                    "attending_raw": attending_raw,
                    "attendance_status": map_attendance(attending_raw),
                }
            )

    return rows
