from datetime import date, datetime, time
from io import BytesIO
import json
import re
from pathlib import Path

from fastapi import Depends, FastAPI, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import text as sqlalchemy_text
from sqlalchemy.orm import Session

from app.database import Base, SessionLocal, engine, get_db
from app.routes.programme import router as programme_router
from app.routes.presence import router as presence_router
from app.models import Camp, ParticipatingGroup, Person, Section, Team, TeamMembership, Task, TaskAssignment, TaskPhase, TaskCategory, Activity, ProgrammeSession, CampRiskAssessment, CampRiskControl, ActivityRiskAssessment, ActivityRiskControl, PresenceWindow
from app.services.osm_event_import import parse_osm_event_attendance_export, normalise_person_match_value

app = FastAPI(title="Camp Command Centre")
app.include_router(programme_router)
app.include_router(presence_router)

BASE_DIR = Path(__file__).resolve().parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


@app.on_event("startup")
def on_startup():
    Base.metadata.create_all(bind=engine)
    ensure_optional_schema_columns()
    normalize_existing_task_statuses()

    with SessionLocal() as db:
        for camp in db.query(Camp).all():
            ensure_default_participating_group(db, camp)
            ensure_default_sections(db, camp)
            ensure_default_task_phases(db, camp)
            ensure_default_task_categories(db, camp)
            apply_default_task_phase_descriptions(db, camp)
            apply_default_task_category_descriptions(db, camp)
            ensure_camp_default_presence_window(db, camp)

def get_latest_camp(db: Session):
    return db.query(Camp).order_by(Camp.start_date.desc()).first()


def camp_start_datetime(camp: Camp):
    return datetime.combine(camp.start_date, camp.start_time or time(18, 0))


def camp_end_datetime(camp: Camp):
    return datetime.combine(camp.end_date, camp.end_time or time(14, 0))


def ensure_camp_default_presence_window(db: Session, camp: Camp):
    # Keep the camp-level default presence rule aligned with camp setup times.
    start_at = camp_start_datetime(camp)
    end_at = camp_end_datetime(camp)

    if end_at <= start_at:
        return None

    default_rule = (
        db.query(PresenceWindow)
        .filter(
            PresenceWindow.camp_id == camp.id,
            PresenceWindow.scope_type == "camp",
            PresenceWindow.scope_id.is_(None),
        )
        .order_by(PresenceWindow.id)
        .first()
    )

    if default_rule is None:
        default_rule = PresenceWindow(
            camp_id=camp.id,
            scope_type="camp",
            scope_id=None,
            starts_at=start_at,
            ends_at=end_at,
            status="Expected",
            notes="Camp default presence window from camp setup.",
        )
        db.add(default_rule)
    else:
        default_rule.starts_at = start_at
        default_rule.ends_at = end_at

        if not default_rule.status:
            default_rule.status = "Expected"

        if not default_rule.notes:
            default_rule.notes = "Camp default presence window from camp setup."

    db.commit()
    return default_rule


def parse_optional_date(value: str):
    value = (value or "").strip()
    if not value:
        return None
    return date.fromisoformat(value)



def get_people_for_activity_leads(db: Session, camp: Camp):
    return (
        db.query(Person)
        .filter(Person.camp_id == camp.id)
        .order_by(Person.person_type, Person.last_name, Person.first_name)
        .all()
    )


def person_display_name(person: Person | None):
    if person is None:
        return "Not set"

    return f"{person.first_name} {person.last_name}"




def normalise_osm_label(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def clean_osm_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def find_osm_value(row_data, group_terms=None, field_terms=None):
    group_terms = [normalise_osm_label(term) for term in (group_terms or [])]
    field_terms = [normalise_osm_label(term) for term in (field_terms or [])]

    for cell in row_data:
        group = normalise_osm_label(cell.get("group", ""))
        field = normalise_osm_label(cell.get("field", ""))
        value = clean_osm_cell(cell.get("value", ""))

        if not value:
            continue

        group_match = True
        field_match = True

        if group_terms:
            group_match = any(term in group for term in group_terms)

        if field_terms:
            field_match = any(term in field for term in field_terms)

        if group_match and field_match:
            return value

    return ""


def combine_names(first_name, last_name):
    parts = [part for part in [clean_osm_cell(first_name), clean_osm_cell(last_name)] if part]
    return " ".join(parts)



def build_osm_duplicate_location_summaries(camp, people, section_lookup, participating_group_lookup):
    summaries = []

    for person in people:
        section = section_lookup.get(person.home_section_id)
        group = participating_group_lookup.get(section.participating_group_id) if section else None

        location_parts = []
        if group:
            location_parts.append(group.name)
        if section:
            location_parts.append(section.name)

        summaries.append(
            {
                "person_id": person.id,
                "display_name": combine_names(person.first_name, person.last_name) or f"Person {person.id}",
                "location": " → ".join(location_parts) if location_parts else "No section assigned",
                "person_url": f"/camps/{camp.id}/people/{person.id}",
                "section_url": f"/camps/{camp.id}/people?open_section_id={section.id}&highlight_person_id={person.id}#person-{person.id}" if section else "",
            }
        )

    return summaries


def build_manual_duplicate_people_data(db: Session, camp: Camp, exclude_person_id: int | None = None):
    sections = (
        db.query(Section)
        .filter(Section.camp_id == camp.id)
        .order_by(Section.sort_order, Section.name)
        .all()
    )
    section_lookup = {section.id: section for section in sections}

    participating_groups = (
        db.query(ParticipatingGroup)
        .filter(ParticipatingGroup.camp_id == camp.id)
        .order_by(ParticipatingGroup.sort_order, ParticipatingGroup.name)
        .all()
    )
    participating_group_lookup = {group.id: group for group in participating_groups}

    people_query = db.query(Person).filter(Person.camp_id == camp.id)

    if exclude_person_id is not None:
        people_query = people_query.filter(Person.id != exclude_person_id)

    people = people_query.order_by(Person.last_name, Person.first_name).all()

    duplicate_people = []

    for person in people:
        section = section_lookup.get(person.home_section_id)
        group = participating_group_lookup.get(section.participating_group_id) if section else None

        location_parts = []
        if group:
            location_parts.append(group.name)
        if section:
            location_parts.append(section.name)

        duplicate_people.append(
            {
                "id": person.id,
                "first_name_key": normalise_person_match_value(person.first_name),
                "last_name_key": normalise_person_match_value(person.last_name),
                "display_name": combine_names(person.first_name, person.last_name) or f"Person {person.id}",
                "home_section_id": person.home_section_id,
                "location": " → ".join(location_parts) if location_parts else "No section assigned",
                "person_url": f"/camps/{camp.id}/people/{person.id}",
            }
        )

    return duplicate_people


def detect_person_type_from_osm_unit(unit_name, default_person_type="Young Person"):
    value = normalise_osm_label(unit_name)

    if "young leader" in value or value == "yl" or "yls" in value:
        return "Young Leader"

    if "leader" in value:
        return "Leader"

    return default_person_type

def extract_osm_candidate(row_data, sheet_name, excel_row):
    member_first_name = (
        find_osm_value(row_data, ["member"], ["first name"])
        or find_osm_value(row_data, [], ["first name"])
    )
    member_last_name = (
        find_osm_value(row_data, ["member"], ["last name"])
        or find_osm_value(row_data, ["member"], ["surname"])
        or find_osm_value(row_data, [], ["last name"])
        or find_osm_value(row_data, [], ["surname"])
    )

    primary_first = find_osm_value(row_data, ["primary contact 1"], ["first name"])
    primary_last = (
        find_osm_value(row_data, ["primary contact 1"], ["last name"])
        or find_osm_value(row_data, ["primary contact 1"], ["surname"])
    )

    emergency_first = find_osm_value(row_data, ["emergency contact"], ["first name"])
    emergency_last = (
        find_osm_value(row_data, ["emergency contact"], ["last name"])
        or find_osm_value(row_data, ["emergency contact"], ["surname"])
    )

    allergies = (
        find_osm_value(row_data, ["essential information"], ["allerg"])
        or find_osm_value(row_data, ["additional information"], ["allerg"])
    )

    medication = (
        find_osm_value(row_data, ["essential information"], ["medication"])
        or find_osm_value(row_data, ["additional information"], ["medication"])
    )

    medical_notes = (
        find_osm_value(row_data, ["essential information"], ["medical"])
        or find_osm_value(row_data, ["additional information"], ["medical"])
    )

    dietary_requirements = (
        find_osm_value(row_data, ["essential information"], ["dietary"])
        or find_osm_value(row_data, ["additional information"], ["dietary"])
    )

    section_unit = (
        find_osm_value(row_data, [], ["six"])
        or find_osm_value(row_data, [], ["patrol"])
        or find_osm_value(row_data, [], ["unit"])
        or find_osm_value(row_data, [], ["section"])
    )

    candidate = {
        "sheet_name": sheet_name,
        "excel_row": excel_row,
        "section_unit": section_unit,
        "first_name": member_first_name,
        "last_name": member_last_name,
        "email": find_osm_value(row_data, ["member"], ["email"]),
        "phone": (
            find_osm_value(row_data, ["member"], ["mobile"])
            or find_osm_value(row_data, ["member"], ["phone"])
        ),
        "primary_contact_name": combine_names(primary_first, primary_last),
        "primary_contact_relationship": find_osm_value(row_data, ["primary contact 1"], ["relationship"]),
        "primary_contact_phone": (
            find_osm_value(row_data, ["primary contact 1"], ["mobile"])
            or find_osm_value(row_data, ["primary contact 1"], ["phone"])
        ),
        "primary_contact_email": find_osm_value(row_data, ["primary contact 1"], ["email"]),
        "emergency_contact_name": combine_names(emergency_first, emergency_last),
        "emergency_contact_relationship": find_osm_value(row_data, ["emergency contact"], ["relationship"]),
        "emergency_contact_phone": (
            find_osm_value(row_data, ["emergency contact"], ["mobile"])
            or find_osm_value(row_data, ["emergency contact"], ["phone"])
        ),
        "emergency_contact_email": find_osm_value(row_data, ["emergency contact"], ["email"]),
        "allergies": allergies,
        "allergy_action": "",
        "medication": medication,
        "medical_notes": medical_notes,
        "dietary_requirements": dietary_requirements,
        "information_source": "OSM Member File",
    }

    candidate["display_name"] = combine_names(candidate["first_name"], candidate["last_name"]) or f"{sheet_name} row {excel_row}"
    candidate["suggested_person_type"] = detect_person_type_from_osm_unit(section_unit)
    candidate["payload"] = json.dumps(candidate)

    return candidate


def parse_osm_member_export(file_bytes):
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    candidates = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        if worksheet.max_row < 3:
            continue

        groups = []
        # OSM exports often leave the first identity columns with a blank
        # group heading. Treat those leading blank-group fields as Member.
        last_group = "Member"

        for col in range(1, worksheet.max_column + 1):
            group_value = clean_osm_cell(worksheet.cell(row=1, column=col).value)
            field_value = clean_osm_cell(worksheet.cell(row=2, column=col).value)

            if group_value:
                last_group = group_value

            groups.append(
                {
                    "group": last_group,
                    "field": field_value,
                    "column": col,
                }
            )

        for row_number in range(3, worksheet.max_row + 1):
            row_data = []
            has_value = False

            for header in groups:
                value = clean_osm_cell(worksheet.cell(row=row_number, column=header["column"]).value)

                if value:
                    has_value = True

                row_data.append(
                    {
                        "group": header["group"],
                        "field": header["field"],
                        "value": value,
                    }
                )

            if not has_value:
                continue

            candidate = extract_osm_candidate(row_data, sheet_name, row_number)

            if candidate["first_name"] or candidate["last_name"]:
                candidates.append(candidate)

    return candidates

def field_is_blank(value):
    return value is None or str(value).strip() == ""



def apply_osm_member_candidate_to_person(person: Person, data: dict, overwrite_existing: bool = False, update_identity: bool = False):
    def incoming_value(field_name):
        return clean_osm_cell(data.get(field_name))

    def current_value(field_name):
        return clean_osm_cell(getattr(person, field_name, ""))

    def apply_field(field_name):
        incoming = incoming_value(field_name)
        existing = current_value(field_name)

        if incoming and (overwrite_existing or not existing):
            setattr(person, field_name, incoming)

    if update_identity:
        if incoming_value("first_name"):
            person.first_name = incoming_value("first_name")

        if incoming_value("last_name"):
            person.last_name = incoming_value("last_name")

        person.is_provisional = False

        if person.role_notes and "Provisional attendee placeholder" in person.role_notes:
            person.role_notes = None

    for field_name in [
        "email",
        "phone",
        "section_unit",
        "primary_contact_name",
        "primary_contact_relationship",
        "primary_contact_phone",
        "primary_contact_email",
        "emergency_contact_name",
        "emergency_contact_relationship",
        "emergency_contact_phone",
        "emergency_contact_email",
        "allergies",
        "allergy_action",
        "medication",
        "medical_notes",
        "dietary_requirements",
    ]:
        apply_field(field_name)

    person.information_source = "OSM Member File"


def get_missing_profile_fields(person: Person):
    missing = []

    if field_is_blank(person.first_name):
        missing.append("First name")

    if field_is_blank(person.last_name):
        missing.append("Last name")

    if field_is_blank(person.person_type):
        missing.append("Person type")

    if person.person_type in ["Young Person", "Young Leader"]:
        checks = [
            ("Home section", person.home_section_id),
            ("Primary contact name", person.primary_contact_name),
            ("Primary contact phone or email", person.primary_contact_phone or person.primary_contact_email),
            ("Emergency contact name", person.emergency_contact_name),
            ("Emergency contact phone", person.emergency_contact_phone),
            ("Allergies", person.allergies),
            ("Medication", person.medication),
            ("Medical notes", person.medical_notes),
            ("Dietary requirements", person.dietary_requirements),
        ]
    elif person.person_type in ["Leader", "Helper"]:
        checks = [
            ("Home section", person.home_section_id),
            ("Phone or email", person.phone or person.email),
            ("Emergency contact name", person.emergency_contact_name),
            ("Emergency contact phone", person.emergency_contact_phone),
            ("Allergies", person.allergies),
            ("Medication", person.medication),
            ("Medical notes", person.medical_notes),
            ("Dietary requirements", person.dietary_requirements),
        ]
    else:
        checks = [
            ("Phone or email", person.phone or person.email),
        ]

    for label, value in checks:
        if field_is_blank(value):
            missing.append(label)

    return missing



def ensure_optional_schema_columns():
    """Small SQLite-safe schema patcher for MVP development.

    SQLAlchemy create_all() creates missing tables, but it does not add new
    columns to tables that already exist. Keep this before any ORM query that
    touches TaskPhase or TaskCategory.
    """
    with engine.begin() as connection:
        for table_name in ["task_phase", "task_category"]:
            columns = [
                row[1]
                for row in connection.execute(
                    sqlalchemy_text(f"PRAGMA table_info({table_name})")
                )
            ]

            if not columns:
                continue

            if "description" not in columns:
                connection.execute(
                    sqlalchemy_text(f"ALTER TABLE {table_name} ADD COLUMN description TEXT")
                )

        optional_columns = [
            ("camp", "start_time", "TIME DEFAULT '18:00:00'"),
            ("camp", "end_time", "TIME DEFAULT '14:00:00'"),
            ("section", "participating_group_id", "INTEGER"),
            ("activity", "badge_notes", "TEXT"),
            ("person", "home_section_id", "INTEGER"),
            ("person", "information_source", "TEXT"),
            ("person", "section_unit", "TEXT"),
            ("person", "dietary_requirements", "TEXT"),
            ("person", "medical_notes", "TEXT"),
            ("person", "medication", "TEXT"),
            ("person", "allergy_action", "TEXT"),
            ("person", "allergies", "TEXT"),
            ("person", "emergency_contact_email", "TEXT"),
            ("person", "emergency_contact_phone", "TEXT"),
            ("person", "emergency_contact_relationship", "TEXT"),
            ("person", "emergency_contact_name", "TEXT"),
            ("person", "primary_contact_email", "TEXT"),
            ("person", "primary_contact_phone", "TEXT"),
            ("person", "primary_contact_relationship", "TEXT"),
            ("person", "primary_contact_name", "TEXT"),
            ("person", "attendance_status", "TEXT"),
            ("person", "is_provisional", "INTEGER DEFAULT 0"),
        ]

        for table_name, column_name, column_type in optional_columns:
            columns = [
                row[1]
                for row in connection.execute(
                    sqlalchemy_text(f"PRAGMA table_info({table_name})")
                )
            ]

            if columns and column_name not in columns:
                connection.execute(
                    sqlalchemy_text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}")
                )



DEFAULT_PARTICIPATING_GROUP_NAME = "Main Group"


def ensure_default_participating_group(db: Session, camp: Camp):
    existing_group = (
        db.query(ParticipatingGroup)
        .filter(ParticipatingGroup.camp_id == camp.id)
        .order_by(ParticipatingGroup.sort_order, ParticipatingGroup.name)
        .first()
    )

    if existing_group is not None:
        return existing_group

    group = ParticipatingGroup(
        camp_id=camp.id,
        name=DEFAULT_PARTICIPATING_GROUP_NAME,
        group_type="Scout Group",
        sort_order=1,
        is_active=True,
    )

    db.add(group)
    db.commit()
    db.refresh(group)

    return group


DEFAULT_SECTIONS = [
    ("Squirrels", "Squirrels"),
    ("Beavers", "Beavers"),
    ("Cubs", "Cubs"),
    ("Scouts", "Scouts"),
    ("Explorers", "Explorers"),
    ("Young Leaders", "Young Leaders"),
    ("Leaders / Adults", "Leaders / Adults"),
    ("Other", "Other"),
]


def ensure_default_sections(db: Session, camp: Camp):
    default_group = ensure_default_participating_group(db, camp)

    existing_sections = (
        db.query(Section)
        .filter(Section.camp_id == camp.id)
        .all()
    )

    if existing_sections:
        changed = False

        for section in existing_sections:
            if section.participating_group_id is None:
                section.participating_group_id = default_group.id
                changed = True

        if changed:
            db.commit()

        return

    for index, (name, section_type) in enumerate(DEFAULT_SECTIONS, start=1):
        db.add(
            Section(
                camp_id=camp.id,
                participating_group_id=default_group.id,
                name=name,
                section_type=section_type,
                sort_order=index,
                is_active=True,
            )
        )

    db.commit()


def get_active_sections(db: Session, camp: Camp):
    return (
        db.query(Section)
        .filter(Section.camp_id == camp.id, Section.is_active == True)
        .order_by(Section.sort_order, Section.name)
        .all()
    )



def get_participating_group_lookup(db: Session, camp: Camp):
    return {
        group.id: group
        for group in db.query(ParticipatingGroup)
        .filter(ParticipatingGroup.camp_id == camp.id)
        .all()
    }



DEFAULT_TASK_CATEGORIES = [
    "Venue",
    "People & Forms",
    "Programme",
    "Equipment",
    "Food",
    "Transport",
    "Documents",
    "Safety / Risk",
    "Communications",
    "Finance",
    "General",
]


DEFAULT_TASK_CATEGORY_DESCRIPTIONS = {
    "Venue": "Site, building, booking, access and venue-related arrangements.",
    "People & Forms": "Young people, adults, permissions, medical information, dietary information and missing responses.",
    "Programme": "Activities, sessions, rotations, activity leads and programme preparation.",
    "Equipment": "Group kit, activity equipment, personal kit checks, storage, loading and returns.",
    "Food": "Menus, food shopping, cooking arrangements, dietary handling and kitchen jobs.",
    "Transport": "Vehicles, drivers, passengers, trailers, loading plans and travel arrangements.",
    "Documents": "Parent packs, leader packs, printed lists, permission records and generated outputs.",
    "Safety / Risk": "Risk assessments, first aid, emergency arrangements and safety checks.",
    "Communications": "Messages to parents, leaders, helpers and other camp contacts.",
    "Finance": "Payments, costs, receipts, reimbursements and budget tracking.",
    "General": "Useful fallback for tasks that do not clearly belong elsewhere.",
}


def ensure_default_task_categories(db: Session, camp: Camp):
    existing_count = (
        db.query(TaskCategory)
        .filter(TaskCategory.camp_id == camp.id)
        .count()
    )

    if existing_count == 0:
        for index, name in enumerate(DEFAULT_TASK_CATEGORIES, start=1):
            db.add(
                TaskCategory(
                    camp_id=camp.id,
                    name=name,
                    description=DEFAULT_TASK_CATEGORY_DESCRIPTIONS.get(name),
                    sort_order=index,
                    is_active=True,
                )
            )

        db.commit()

    existing_names = {
        row.name
        for row in db.query(TaskCategory)
        .filter(TaskCategory.camp_id == camp.id)
        .all()
    }

    task_category_names = {
        task.category.strip()
        for task in db.query(Task)
        .filter(Task.camp_id == camp.id, Task.category.isnot(None))
        .all()
        if task.category and task.category.strip()
    }

    next_order = (
        db.query(TaskCategory)
        .filter(TaskCategory.camp_id == camp.id)
        .count()
        + 1
    )

    changed = False

    for name in sorted(task_category_names):
        if name not in existing_names:
            db.add(
                TaskCategory(
                    camp_id=camp.id,
                    name=name,
                    description=DEFAULT_TASK_CATEGORY_DESCRIPTIONS.get(name),
                    sort_order=next_order,
                    is_active=True,
                )
            )
            next_order += 1
            changed = True

    if changed:
        db.commit()


def get_active_task_categories(db: Session, camp: Camp):
    return (
        db.query(TaskCategory)
        .filter(TaskCategory.camp_id == camp.id, TaskCategory.is_active == True)
        .order_by(TaskCategory.sort_order, TaskCategory.name)
        .all()
    )


def apply_default_task_category_descriptions(db: Session, camp: Camp):
    changed = False

    for category in db.query(TaskCategory).filter(TaskCategory.camp_id == camp.id).all():
        if not category.description and category.name in DEFAULT_TASK_CATEGORY_DESCRIPTIONS:
            category.description = DEFAULT_TASK_CATEGORY_DESCRIPTIONS[category.name]
            changed = True

    if changed:
        db.commit()


DEFAULT_TASK_PHASES = [
    "Early Planning",
    "Preparation",
    "Final Week",
    "Camp Setup",
    "During Camp",
    "Pack Down",
    "After Camp",
]


def ensure_default_task_phases(db: Session, camp: Camp):
    existing_count = (
        db.query(TaskPhase)
        .filter(TaskPhase.camp_id == camp.id)
        .count()
    )

    if existing_count > 0:
        return

    for index, name in enumerate(DEFAULT_TASK_PHASES, start=1):
        db.add(
            TaskPhase(
                camp_id=camp.id,
                name=name,
                sort_order=index,
                is_active=True,
            )
        )

    db.commit()


def get_active_task_phases(db: Session, camp: Camp):
    return (
        db.query(TaskPhase)
        .filter(TaskPhase.camp_id == camp.id, TaskPhase.is_active == True)
        .order_by(TaskPhase.sort_order, TaskPhase.name)
        .all()
    )


DEFAULT_TASK_PHASE_DESCRIPTIONS = {
    "Early Planning": "Early decisions and bookings, usually before detailed camp preparation begins.",
    "Preparation": "General preparation work before the final week, including organising people, tasks and materials.",
    "Final Week": "Tasks that must be completed in the final week before camp.",
    "Camp Setup": "Work needed to set up the site or venue before activities begin.",
    "During Camp": "Tasks that happen while camp is running.",
    "Pack Down": "Jobs for packing away, clearing the site and checking nothing is left behind.",
    "After Camp": "Follow-up work after camp, such as returns, reviews, payments and lessons learned.",
}

DEFAULT_TASK_CATEGORY_DESCRIPTIONS = {
    "Venue": "Site, building, booking, access and venue-related arrangements.",
    "People & Forms": "Young people, adults, permissions, medical information, dietary information and missing responses.",
    "Programme": "Activities, sessions, rotations, activity leads and programme preparation.",
    "Equipment": "Group kit, activity equipment, personal kit checks, storage, loading and returns.",
    "Food": "Menus, food shopping, cooking arrangements, dietary handling and kitchen jobs.",
    "Transport": "Vehicles, drivers, passengers, trailers, loading plans and travel arrangements.",
    "Documents": "Parent packs, leader packs, printed lists, permission records and generated outputs.",
    "Safety / Risk": "Risk assessments, first aid, emergency arrangements and safety checks.",
    "Communications": "Messages to parents, leaders, helpers and other camp contacts.",
    "Finance": "Payments, costs, receipts, reimbursements and budget tracking.",
    "General": "Useful fallback for tasks that do not clearly belong elsewhere.",
}


def apply_default_task_phase_descriptions(db: Session, camp: Camp):
    changed = False

    for phase in db.query(TaskPhase).filter(TaskPhase.camp_id == camp.id).all():
        if not phase.description and phase.name in DEFAULT_TASK_PHASE_DESCRIPTIONS:
            phase.description = DEFAULT_TASK_PHASE_DESCRIPTIONS[phase.name]
            changed = True

    if changed:
        db.commit()


def apply_default_task_category_descriptions(db: Session, camp: Camp):
    changed = False

    for category in db.query(TaskCategory).filter(TaskCategory.camp_id == camp.id).all():
        if not category.description and category.name in DEFAULT_TASK_CATEGORY_DESCRIPTIONS:
            category.description = DEFAULT_TASK_CATEGORY_DESCRIPTIONS[category.name]
            changed = True

    if changed:
        db.commit()


RISK_STATUSES = ["Not Started", "Draft", "Ready for Review", "Submitted", "Approved", "Needs Update"]
ACTIVITY_RISK_SOURCE_TYPES = ["Created in app", "External provider", "Existing/generic assessment"]


def ensure_camp_risk_assessment(db: Session, camp: Camp):
    risk_assessment = (
        db.query(CampRiskAssessment)
        .filter(CampRiskAssessment.camp_id == camp.id)
        .first()
    )

    if risk_assessment is None:
        risk_assessment = CampRiskAssessment(
            camp_id=camp.id,
            title=f"{camp.name} Risk Assessment",
            status="Not Started",
        )
        db.add(risk_assessment)
        db.commit()
        db.refresh(risk_assessment)

    return risk_assessment


def ensure_activity_risk_assessment(db: Session, camp: Camp, activity: Activity):
    risk_assessment = (
        db.query(ActivityRiskAssessment)
        .filter(
            ActivityRiskAssessment.camp_id == camp.id,
            ActivityRiskAssessment.activity_id == activity.id,
        )
        .first()
    )

    if risk_assessment is None:
        risk_assessment = ActivityRiskAssessment(
            camp_id=camp.id,
            activity_id=activity.id,
            source_type="Created in app",
            status="Not Started",
        )
        db.add(risk_assessment)
        db.commit()
        db.refresh(risk_assessment)

    return risk_assessment


TASK_STATUSES = ["To Do", "In Progress", "Blocked", "Done"]

OLD_TASK_STATUS_MAP = {
    "Draft": "To Do",
    "Planned": "To Do",
    "Unassigned": "To Do",
    "Assigned": "To Do",
    "Accepted": "To Do",
    "Ready for Check": "In Progress",
    "Complete": "Done",
    "Checked": "Done",
    "Cancelled": "Done",
    "Not Needed": "Done",
}


def normalize_existing_task_statuses():
    with SessionLocal() as db:
        changed = False

        for task in db.query(Task).all():
            if task.status in OLD_TASK_STATUS_MAP:
                task.status = OLD_TASK_STATUS_MAP[task.status]
                changed = True

        if changed:
            db.commit()


def update_task_status_from_assignments(db: Session, task: Task):
    # Assignment ownership is derived from TaskAssignment rows.
    # Do not change task.status when assignments are added or removed.
    return
def task_assignment_duplicate_exists(
    db: Session,
    task: Task,
    assigned_person_id: int | None = None,
    assigned_team_id: int | None = None,
    exclude_assignment_id: int | None = None,
):
    query = db.query(TaskAssignment).filter(
        TaskAssignment.camp_id == task.camp_id,
        TaskAssignment.task_id == task.id,
    )

    if assigned_person_id is not None:
        query = query.filter(TaskAssignment.assigned_person_id == assigned_person_id)

    if assigned_team_id is not None:
        query = query.filter(TaskAssignment.assigned_team_id == assigned_team_id)

    if exclude_assignment_id is not None:
        query = query.filter(TaskAssignment.id != exclude_assignment_id)

    return query.first() is not None


def get_available_task_assignees(db: Session, camp: Camp, task: Task):
    existing_person_ids = [
        row.assigned_person_id
        for row in db.query(TaskAssignment)
        .filter(
            TaskAssignment.camp_id == camp.id,
            TaskAssignment.task_id == task.id,
            TaskAssignment.assigned_person_id.isnot(None),
        )
        .all()
    ]

    existing_team_ids = [
        row.assigned_team_id
        for row in db.query(TaskAssignment)
        .filter(
            TaskAssignment.camp_id == camp.id,
            TaskAssignment.task_id == task.id,
            TaskAssignment.assigned_team_id.isnot(None),
        )
        .all()
    ]

    people_query = db.query(Person).filter(Person.camp_id == camp.id)
    teams_query = db.query(Team).filter(Team.camp_id == camp.id)

    if existing_person_ids:
        people_query = people_query.filter(Person.id.notin_(existing_person_ids))

    if existing_team_ids:
        teams_query = teams_query.filter(Team.id.notin_(existing_team_ids))

    people = people_query.order_by(Person.person_type, Person.last_name, Person.first_name).all()
    teams = teams_query.order_by(Team.team_type, Team.name).all()

    return people, teams


@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, db: Session = Depends(get_db)):
    current_camp = get_latest_camp(db)

    people_count = 0
    task_count = 0
    activity_count = 0
    programme_session_count = 0

    if current_camp:
        people_count = db.query(Person).filter(Person.camp_id == current_camp.id).count()
        task_count = db.query(Task).filter(Task.camp_id == current_camp.id).count()
        activity_count = db.query(Activity).filter(Activity.camp_id == current_camp.id).count()
        programme_session_count = db.query(ProgrammeSession).filter(ProgrammeSession.camp_id == current_camp.id).count()

    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "camp": current_camp,
            "people_count": people_count,
            "task_count": task_count,
            "activity_count": activity_count,
            "programme_sessions": programme_session_count,
            "readiness": 0,
        },
    )


@app.get("/camps", response_class=HTMLResponse)
async def camp_list(request: Request, db: Session = Depends(get_db)):
    camps = db.query(Camp).order_by(Camp.start_date.desc()).all()

    return templates.TemplateResponse(
        "camps/list.html",
        {
            "request": request,
            "camps": camps,
        },
    )


@app.get("/camps/new", response_class=HTMLResponse)
async def new_camp_form(request: Request):
    return templates.TemplateResponse(
        "camps/new.html",
        {
            "request": request,
            "error": None,
        },
    )


@app.post("/camps/new")
async def create_camp(
    request: Request,
    name: str = Form(...),
    camp_type: str = Form("Campsite Camp"),
    start_date: date = Form(...),
    start_time: time = Form(time(18, 0)),
    end_date: date = Form(...),
    end_time: time = Form(time(14, 0)),
    venue_name: str = Form(""),
    camp_leader: str = Form(""),
    permit_holder: str = Form(""),
    notes: str = Form(""),
    db: Session = Depends(get_db),
):
    if datetime.combine(end_date, end_time) <= datetime.combine(start_date, start_time):
        return templates.TemplateResponse(
            "camps/new.html",
            {
                "request": request,
                "error": "The camp end date/time must be after the start date/time.",
            },
            status_code=400,
        )

    camp = Camp(
        name=name.strip(),
        camp_type=camp_type,
        start_date=start_date,
        start_time=start_time,
        end_date=end_date,
        end_time=end_time,
        venue_name=venue_name.strip() or None,
        camp_leader=camp_leader.strip() or None,
        permit_holder=permit_holder.strip() or None,
        status="Planning",
        notes=notes.strip() or None,
    )

    db.add(camp)
    db.commit()
    db.refresh(camp)

    ensure_default_participating_group(db, camp)
    ensure_default_sections(db, camp)
    ensure_default_task_phases(db, camp)
    ensure_default_task_categories(db, camp)
    apply_default_task_phase_descriptions(db, camp)
    apply_default_task_category_descriptions(db, camp)
    ensure_camp_default_presence_window(db, camp)

    return RedirectResponse(url=f"/camps/{camp.id}", status_code=303)


@app.get("/camps/{camp_id}/edit", response_class=HTMLResponse)
async def edit_camp_form(request: Request, camp_id: int, db: Session = Depends(get_db)):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Camp not found.",
            },
            status_code=404,
        )

    return templates.TemplateResponse(
        "camps/edit.html",
        {
            "request": request,
            "camp": camp,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/edit")
async def update_camp(
    request: Request,
    camp_id: int,
    name: str = Form(...),
    camp_type: str = Form("Campsite Camp"),
    start_date: date = Form(...),
    start_time: time = Form(time(18, 0)),
    end_date: date = Form(...),
    end_time: time = Form(time(14, 0)),
    venue_name: str = Form(""),
    camp_leader: str = Form(""),
    permit_holder: str = Form(""),
    status: str = Form("Planning"),
    notes: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Camp not found.",
            },
            status_code=404,
        )

    if datetime.combine(end_date, end_time) <= datetime.combine(start_date, start_time):
        return templates.TemplateResponse(
            "camps/edit.html",
            {
                "request": request,
                "camp": camp,
                "error": "The camp end date/time must be after the start date/time.",
            },
            status_code=400,
        )

    camp.name = name.strip()
    camp.camp_type = camp_type
    camp.start_date = start_date
    camp.start_time = start_time
    camp.end_date = end_date
    camp.end_time = end_time
    camp.venue_name = venue_name.strip() or None
    camp.camp_leader = camp_leader.strip() or None
    camp.permit_holder = permit_holder.strip() or None
    camp.status = status
    camp.notes = notes.strip() or None

    db.commit()
    ensure_camp_default_presence_window(db, camp)

    return RedirectResponse(url=f"/camps/{camp.id}", status_code=303)





@app.get("/camps/{camp_id}/delete", response_class=HTMLResponse)
async def delete_camp_form(
    request: Request,
    camp_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    return templates.TemplateResponse(
        "camps/delete.html",
        {
            "request": request,
            "camp": camp,
        },
    )


@app.post("/camps/{camp_id}/delete")
async def delete_camp(
    request: Request,
    camp_id: int,
    confirm_delete_camp: str | None = Form(None),
    confirm_camp_name: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return RedirectResponse(url="/camps?camp_error=not_found", status_code=303)

    if confirm_delete_camp != "yes":
        return RedirectResponse(
            url=f"/camps/{camp.id}/delete?camp_error=delete_box_not_ticked",
            status_code=303,
        )

    if confirm_camp_name.strip() != camp.name:
        return RedirectResponse(
            url=f"/camps/{camp.id}/delete?camp_error=delete_name_mismatch",
            status_code=303,
        )

    camp_name = camp.name

    camp_ra_ids = [
        row.id
        for row in db.query(CampRiskAssessment)
        .filter(CampRiskAssessment.camp_id == camp.id)
        .all()
    ]

    activity_ra_ids = [
        row.id
        for row in db.query(ActivityRiskAssessment)
        .filter(ActivityRiskAssessment.camp_id == camp.id)
        .all()
    ]

    if camp_ra_ids:
        db.query(CampRiskControl).filter(
            CampRiskControl.risk_assessment_id.in_(camp_ra_ids)
        ).delete(synchronize_session=False)

    if activity_ra_ids:
        db.query(ActivityRiskControl).filter(
            ActivityRiskControl.risk_assessment_id.in_(activity_ra_ids)
        ).delete(synchronize_session=False)

    db.query(CampRiskAssessment).filter(CampRiskAssessment.camp_id == camp.id).delete(synchronize_session=False)
    db.query(ActivityRiskAssessment).filter(ActivityRiskAssessment.camp_id == camp.id).delete(synchronize_session=False)

    db.query(ProgrammeSession).filter(ProgrammeSession.camp_id == camp.id).delete(synchronize_session=False)

    db.query(TaskAssignment).filter(TaskAssignment.camp_id == camp.id).delete(synchronize_session=False)
    db.query(Task).filter(Task.camp_id == camp.id).delete(synchronize_session=False)
    db.query(TaskPhase).filter(TaskPhase.camp_id == camp.id).delete(synchronize_session=False)
    db.query(TaskCategory).filter(TaskCategory.camp_id == camp.id).delete(synchronize_session=False)

    db.query(TeamMembership).filter(TeamMembership.camp_id == camp.id).delete(synchronize_session=False)
    db.query(Team).filter(Team.camp_id == camp.id).delete(synchronize_session=False)

    db.query(Activity).filter(Activity.camp_id == camp.id).delete(synchronize_session=False)

    db.query(Person).filter(Person.camp_id == camp.id).delete(synchronize_session=False)
    db.query(Section).filter(Section.camp_id == camp.id).delete(synchronize_session=False)
    db.query(ParticipatingGroup).filter(ParticipatingGroup.camp_id == camp.id).delete(synchronize_session=False)

    db.delete(camp)
    db.commit()

    return RedirectResponse(
        url=f"/camps?camp_deleted={camp_name}",
        status_code=303,
    )


@app.get("/camps/{camp_id}", response_class=HTMLResponse)
async def camp_detail(request: Request, camp_id: int, db: Session = Depends(get_db)):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Camp not found.",
            },
            status_code=404,
        )

    people_count = db.query(Person).filter(Person.camp_id == camp.id).count()
    team_count = db.query(Team).filter(Team.camp_id == camp.id).count()
    task_count = db.query(Task).filter(Task.camp_id == camp.id).count()
    activity_count = db.query(Activity).filter(Activity.camp_id == camp.id).count()
    programme_session_count = db.query(ProgrammeSession).filter(ProgrammeSession.camp_id == camp.id).count()

    return templates.TemplateResponse(
        "camps/detail.html",
        {
            "request": request,
            "camp": camp,
            "people_count": people_count,
            "team_count": team_count,
            "task_count": task_count,
            "activity_count": activity_count,
            "programme_sessions": programme_session_count,
            "readiness": 0,
        },
    )



PARTICIPATING_GROUP_TYPES = [
    "Scout Group",
    "Explorer Unit",
    "District",
    "Visiting Group",
    "Partner Organisation",
    "Other",
]


@app.get("/camps/{camp_id}/participating-groups", response_class=HTMLResponse)
async def participating_group_list(request: Request, camp_id: int, db: Session = Depends(get_db)):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    ensure_default_participating_group(db, camp)
    ensure_default_sections(db, camp)

    groups = (
        db.query(ParticipatingGroup)
        .filter(ParticipatingGroup.camp_id == camp.id)
        .order_by(ParticipatingGroup.sort_order, ParticipatingGroup.name)
        .all()
    )

    group_section_counts = {
        group.id: db.query(Section)
        .filter(Section.camp_id == camp.id, Section.participating_group_id == group.id)
        .count()
        for group in groups
    }

    group_person_counts = {
        group.id: db.query(Person)
        .join(Section, Person.home_section_id == Section.id)
        .filter(
            Person.camp_id == camp.id,
            Section.camp_id == camp.id,
            Section.participating_group_id == group.id,
        )
        .count()
        for group in groups
    }

    return templates.TemplateResponse(
        "participating_groups/list.html",
        {
            "request": request,
            "camp": camp,
            "groups": groups,
            "group_section_counts": group_section_counts,
            "group_person_counts": group_person_counts,
        },
    )


@app.get("/camps/{camp_id}/participating-groups/new", response_class=HTMLResponse)
async def new_participating_group_form(request: Request, camp_id: int, db: Session = Depends(get_db)):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    return templates.TemplateResponse(
        "participating_groups/new.html",
        {
            "request": request,
            "camp": camp,
            "group_types": PARTICIPATING_GROUP_TYPES,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/participating-groups/new")
async def create_participating_group(
    request: Request,
    camp_id: int,
    name: str = Form(...),
    group_type: str = Form("Scout Group"),
    contact_name: str = Form(""),
    contact_email: str = Form(""),
    contact_phone: str = Form(""),
    notes: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    clean_name = name.strip()

    if not clean_name:
        return templates.TemplateResponse(
            "participating_groups/new.html",
            {
                "request": request,
                "camp": camp,
                "group_types": PARTICIPATING_GROUP_TYPES,
                "error": "Participating group name is required.",
            },
            status_code=400,
        )

    next_order = (
        db.query(ParticipatingGroup)
        .filter(ParticipatingGroup.camp_id == camp.id)
        .count()
        + 1
    )

    group = ParticipatingGroup(
        camp_id=camp.id,
        name=clean_name,
        group_type=group_type,
        contact_name=contact_name.strip() or None,
        contact_email=contact_email.strip() or None,
        contact_phone=contact_phone.strip() or None,
        notes=notes.strip() or None,
        sort_order=next_order,
        is_active=True,
    )

    db.add(group)
    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/participating-groups", status_code=303)



@app.post("/camps/{camp_id}/participating-groups/{group_id}/delete")
async def delete_participating_group(
    request: Request,
    camp_id: int,
    group_id: int,
    confirm_delete_group: str | None = Form(None),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    group = (
        db.query(ParticipatingGroup)
        .filter(
            ParticipatingGroup.id == group_id,
            ParticipatingGroup.camp_id == camp.id,
        )
        .first()
    )

    if group is None:
        return RedirectResponse(
            url=f"/camps/{camp.id}/participating-groups?group_error=not_found",
            status_code=303,
        )

    if confirm_delete_group != "yes":
        return RedirectResponse(
            url=f"/camps/{camp.id}/participating-groups?group_error=delete_not_confirmed",
            status_code=303,
        )

    section_count = (
        db.query(Section)
        .filter(
            Section.camp_id == camp.id,
            Section.participating_group_id == group.id,
        )
        .count()
    )

    if section_count > 0:
        return RedirectResponse(
            url=f"/camps/{camp.id}/participating-groups?group_error=group_not_empty",
            status_code=303,
        )

    db.delete(group)
    db.commit()

    return RedirectResponse(
        url=f"/camps/{camp.id}/participating-groups?group_deleted=1",
        status_code=303,
    )



@app.post("/camps/{camp_id}/participating-groups/bulk-delete")
async def bulk_delete_participating_groups(
    request: Request,
    camp_id: int,
    selected_group_id: list[int] | None = Form(None),
    confirm_delete_groups: str | None = Form(None),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    group_ids = selected_group_id or []

    if not group_ids:
        return RedirectResponse(
            url=f"/camps/{camp.id}/participating-groups?bulk_error=no_selection#group-bulk-actions",
            status_code=303,
        )

    if confirm_delete_groups != "yes":
        return RedirectResponse(
            url=f"/camps/{camp.id}/participating-groups?bulk_error=delete_not_confirmed#group-bulk-actions",
            status_code=303,
        )

    groups = (
        db.query(ParticipatingGroup)
        .filter(
            ParticipatingGroup.camp_id == camp.id,
            ParticipatingGroup.id.in_(group_ids),
        )
        .all()
    )

    if not groups:
        return RedirectResponse(
            url=f"/camps/{camp.id}/participating-groups?bulk_error=no_valid_groups#group-bulk-actions",
            status_code=303,
        )

    real_group_ids = [group.id for group in groups]

    section_count = (
        db.query(Section)
        .filter(
            Section.camp_id == camp.id,
            Section.participating_group_id.in_(real_group_ids),
        )
        .count()
    )

    if section_count > 0:
        return RedirectResponse(
            url=f"/camps/{camp.id}/participating-groups?bulk_error=groups_not_empty#group-bulk-actions",
            status_code=303,
        )

    deleted_count = len(groups)

    for group in groups:
        db.delete(group)

    db.commit()

    return RedirectResponse(
        url=f"/camps/{camp.id}/participating-groups?bulk_deleted={deleted_count}#group-bulk-actions",
        status_code=303,
    )


@app.get("/camps/{camp_id}/participating-groups/{group_id}/edit", response_class=HTMLResponse)
async def edit_participating_group_form(
    request: Request,
    camp_id: int,
    group_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    group = (
        db.query(ParticipatingGroup)
        .filter(ParticipatingGroup.id == group_id, ParticipatingGroup.camp_id == camp.id)
        .first()
    )

    if group is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Participating group not found."},
            status_code=404,
        )

    return templates.TemplateResponse(
        "participating_groups/edit.html",
        {
            "request": request,
            "camp": camp,
            "group": group,
            "group_types": PARTICIPATING_GROUP_TYPES,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/participating-groups/{group_id}/edit")
async def update_participating_group(
    request: Request,
    camp_id: int,
    group_id: int,
    name: str = Form(...),
    group_type: str = Form("Scout Group"),
    contact_name: str = Form(""),
    contact_email: str = Form(""),
    contact_phone: str = Form(""),
    notes: str = Form(""),
    sort_order: int = Form(0),
    is_active: str | None = Form(None),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    group = (
        db.query(ParticipatingGroup)
        .filter(ParticipatingGroup.id == group_id, ParticipatingGroup.camp_id == camp.id)
        .first()
    )

    if group is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Participating group not found."},
            status_code=404,
        )

    clean_name = name.strip()

    if not clean_name:
        return templates.TemplateResponse(
            "participating_groups/edit.html",
            {
                "request": request,
                "camp": camp,
                "group": group,
                "group_types": PARTICIPATING_GROUP_TYPES,
                "error": "Participating group name is required.",
            },
            status_code=400,
        )

    group.name = clean_name
    group.group_type = group_type
    group.contact_name = contact_name.strip() or None
    group.contact_email = contact_email.strip() or None
    group.contact_phone = contact_phone.strip() or None
    group.notes = notes.strip() or None
    group.sort_order = sort_order
    group.is_active = is_active == "yes"

    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/participating-groups", status_code=303)


@app.get("/camps/{camp_id}/sections", response_class=HTMLResponse)
async def camp_section_list(request: Request, camp_id: int, db: Session = Depends(get_db)):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    ensure_default_participating_group(db, camp)
    ensure_default_sections(db, camp)

    participating_groups = (
        db.query(ParticipatingGroup)
        .filter(ParticipatingGroup.camp_id == camp.id)
        .order_by(ParticipatingGroup.sort_order, ParticipatingGroup.name)
        .all()
    )
    participating_group_lookup = {group.id: group for group in participating_groups}

    sections = (
        db.query(Section)
        .filter(Section.camp_id == camp.id)
        .order_by(Section.sort_order, Section.name)
        .all()
    )

    section_person_counts = {
        section.id: db.query(Person)
        .filter(Person.camp_id == camp.id, Person.home_section_id == section.id)
        .count()
        for section in sections
    }

    return templates.TemplateResponse(
        "sections/list.html",
        {
            "request": request,
            "camp": camp,
            "sections": sections,
            "participating_groups": participating_groups,
            "participating_group_lookup": participating_group_lookup,
            "section_person_counts": section_person_counts,
        },
    )



@app.post("/camps/{camp_id}/sections/bulk-move")
async def bulk_move_sections(
    request: Request,
    camp_id: int,
    selected_section_id: list[int] | None = Form(None),
    target_participating_group_id: int = Form(...),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    section_ids = selected_section_id or []

    if not section_ids:
        return RedirectResponse(
            url=f"/camps/{camp.id}/sections?bulk_error=no_selection#section-bulk-actions",
            status_code=303,
        )

    target_group = (
        db.query(ParticipatingGroup)
        .filter(
            ParticipatingGroup.id == target_participating_group_id,
            ParticipatingGroup.camp_id == camp.id,
        )
        .first()
    )

    if target_group is None:
        return RedirectResponse(
            url=f"/camps/{camp.id}/sections?bulk_error=invalid_group#section-bulk-actions",
            status_code=303,
        )

    sections = (
        db.query(Section)
        .filter(
            Section.camp_id == camp.id,
            Section.id.in_(section_ids),
        )
        .all()
    )

    if not sections:
        return RedirectResponse(
            url=f"/camps/{camp.id}/sections?bulk_error=no_valid_sections#section-bulk-actions",
            status_code=303,
        )

    moved = 0

    for section in sections:
        section.participating_group_id = target_group.id
        moved += 1

    db.commit()

    return RedirectResponse(
        url=f"/camps/{camp.id}/sections?bulk_moved={moved}#section-bulk-actions",
        status_code=303,
    )



@app.post("/camps/{camp_id}/sections/bulk-delete")
async def bulk_delete_sections(
    request: Request,
    camp_id: int,
    selected_section_id: list[int] | None = Form(None),
    confirm_delete_sections: str | None = Form(None),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    section_ids = selected_section_id or []

    if not section_ids:
        return RedirectResponse(
            url=f"/camps/{camp.id}/sections?bulk_error=no_selection#section-bulk-actions",
            status_code=303,
        )

    if confirm_delete_sections != "yes":
        return RedirectResponse(
            url=f"/camps/{camp.id}/sections?bulk_error=delete_not_confirmed#section-bulk-actions",
            status_code=303,
        )

    sections = (
        db.query(Section)
        .filter(
            Section.camp_id == camp.id,
            Section.id.in_(section_ids),
        )
        .all()
    )

    if not sections:
        return RedirectResponse(
            url=f"/camps/{camp.id}/sections?bulk_error=no_valid_sections#section-bulk-actions",
            status_code=303,
        )

    section_ids_to_delete = [section.id for section in sections]

    people_count = (
        db.query(Person)
        .filter(
            Person.camp_id == camp.id,
            Person.home_section_id.in_(section_ids_to_delete),
        )
        .count()
    )

    if people_count > 0:
        return RedirectResponse(
            url=f"/camps/{camp.id}/sections?bulk_error=sections_not_empty#section-bulk-actions",
            status_code=303,
        )

    deleted_count = len(sections)

    for section in sections:
        db.delete(section)

    db.commit()

    return RedirectResponse(
        url=f"/camps/{camp.id}/sections?bulk_deleted={deleted_count}#section-bulk-actions",
        status_code=303,
    )


@app.get("/camps/{camp_id}/sections/new", response_class=HTMLResponse)
async def new_section_form(request: Request, camp_id: int, db: Session = Depends(get_db)):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    participating_groups = (
        db.query(ParticipatingGroup)
        .filter(ParticipatingGroup.camp_id == camp.id, ParticipatingGroup.is_active == True)
        .order_by(ParticipatingGroup.sort_order, ParticipatingGroup.name)
        .all()
    )

    return templates.TemplateResponse(
        "sections/new.html",
        {
            "request": request,
            "camp": camp,
            "participating_groups": participating_groups,
            "section_types": [section_type for name, section_type in DEFAULT_SECTIONS],
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/sections/new")
async def create_section(
    request: Request,
    camp_id: int,
    name: str = Form(...),
    participating_group_id: int = Form(...),
    section_type: str = Form("Other"),
    notes: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    participating_groups = (
        db.query(ParticipatingGroup)
        .filter(ParticipatingGroup.camp_id == camp.id, ParticipatingGroup.is_active == True)
        .order_by(ParticipatingGroup.sort_order, ParticipatingGroup.name)
        .all()
    )

    selected_group = (
        db.query(ParticipatingGroup)
        .filter(
            ParticipatingGroup.id == participating_group_id,
            ParticipatingGroup.camp_id == camp.id,
        )
        .first()
    )

    clean_name = name.strip()

    if selected_group is None:
        return templates.TemplateResponse(
            "sections/new.html",
            {
                "request": request,
                "camp": camp,
                "participating_groups": participating_groups,
                "section_types": [section_type for name, section_type in DEFAULT_SECTIONS],
                "error": "Please choose a valid participating group.",
            },
            status_code=400,
        )

    if not clean_name:
        return templates.TemplateResponse(
            "sections/new.html",
            {
                "request": request,
                "camp": camp,
                "participating_groups": participating_groups,
                "section_types": [section_type for name, section_type in DEFAULT_SECTIONS],
                "error": "Section name is required.",
            },
            status_code=400,
        )

    next_order = (
        db.query(Section)
        .filter(Section.camp_id == camp.id)
        .count()
        + 1
    )

    section = Section(
        camp_id=camp.id,
        name=clean_name,
        section_type=section_type,
        participating_group_id=selected_group.id,
        notes=notes.strip() or None,
        sort_order=next_order,
        is_active=True,
    )

    db.add(section)
    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/sections", status_code=303)


@app.get("/camps/{camp_id}/sections/{section_id}/edit", response_class=HTMLResponse)
async def edit_section_form(
    request: Request,
    camp_id: int,
    section_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    section = (
        db.query(Section)
        .filter(Section.id == section_id, Section.camp_id == camp.id)
        .first()
    )

    if section is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Section not found."},
            status_code=404,
        )

    participating_groups = (
        db.query(ParticipatingGroup)
        .filter(ParticipatingGroup.camp_id == camp.id, ParticipatingGroup.is_active == True)
        .order_by(ParticipatingGroup.sort_order, ParticipatingGroup.name)
        .all()
    )

    return templates.TemplateResponse(
        "sections/edit.html",
        {
            "request": request,
            "camp": camp,
            "section": section,
            "participating_groups": participating_groups,
            "section_types": [section_type for name, section_type in DEFAULT_SECTIONS],
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/sections/{section_id}/edit")
async def update_section(
    request: Request,
    camp_id: int,
    section_id: int,
    name: str = Form(...),
    participating_group_id: int = Form(...),
    section_type: str = Form("Other"),
    notes: str = Form(""),
    sort_order: int = Form(0),
    is_active: str | None = Form(None),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    section = (
        db.query(Section)
        .filter(Section.id == section_id, Section.camp_id == camp.id)
        .first()
    )

    if section is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Section not found."},
            status_code=404,
        )

    participating_groups = (
        db.query(ParticipatingGroup)
        .filter(
            ParticipatingGroup.camp_id == camp.id,
            ParticipatingGroup.is_active == True,
        )
        .order_by(ParticipatingGroup.sort_order, ParticipatingGroup.name)
        .all()
    )

    selected_group = (
        db.query(ParticipatingGroup)
        .filter(
            ParticipatingGroup.id == participating_group_id,
            ParticipatingGroup.camp_id == camp.id,
        )
        .first()
    )

    if selected_group is None:
        return templates.TemplateResponse(
            "sections/edit.html",
            {
                "request": request,
                "camp": camp,
                "section": section,
                "participating_groups": participating_groups,
                "section_types": [section_type for name, section_type in DEFAULT_SECTIONS],
                "error": "Please choose a valid participating group.",
            },
            status_code=400,
        )

    clean_name = name.strip()

    if not clean_name:
        return templates.TemplateResponse(
            "sections/edit.html",
            {
                "request": request,
                "camp": camp,
                "section": section,
                "participating_groups": participating_groups,
                "section_types": [section_type for name, section_type in DEFAULT_SECTIONS],
                "error": "Section name is required.",
            },
            status_code=400,
        )

    section.name = clean_name
    section.participating_group_id = selected_group.id
    section.section_type = section_type
    section.notes = notes.strip() or None
    section.sort_order = sort_order
    section.is_active = is_active == "yes"

    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/sections", status_code=303)


@app.get("/people", response_class=HTMLResponse)
async def people_page(request: Request, db: Session = Depends(get_db)):
    camp = get_latest_camp(db)

    if camp is None:
        return templates.TemplateResponse("people.html", {"request": request})

    return RedirectResponse(url=f"/camps/{camp.id}/people", status_code=303)


@app.get("/camps/{camp_id}/people", response_class=HTMLResponse)
async def camp_people_list(
    request: Request,
    camp_id: int,
    open_section_id: int | None = None,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Camp not found.",
            },
            status_code=404,
        )

    ensure_default_participating_group(db, camp)
    ensure_default_sections(db, camp)

    people = (
        db.query(Person)
        .filter(Person.camp_id == camp.id)
        .order_by(Person.person_type, Person.last_name, Person.first_name)
        .all()
    )

    participating_groups = (
        db.query(ParticipatingGroup)
        .filter(ParticipatingGroup.camp_id == camp.id)
        .order_by(ParticipatingGroup.sort_order, ParticipatingGroup.name)
        .all()
    )

    sections = (
        db.query(Section)
        .filter(Section.camp_id == camp.id)
        .order_by(Section.sort_order, Section.name)
        .all()
    )

    participating_group_lookup = {group.id: group for group in participating_groups}
    section_lookup = {section.id: section for section in sections}

    sections_by_group = {group.id: [] for group in participating_groups}
    sections_without_group = []

    for section in sections:
        if section.participating_group_id in sections_by_group:
            sections_by_group[section.participating_group_id].append(section)
        else:
            sections_without_group.append(section)

    people_by_section = {section.id: [] for section in sections}
    people_without_section = []

    for person in people:
        if person.home_section_id in people_by_section:
            people_by_section[person.home_section_id].append(person)
        else:
            people_without_section.append(person)

    person_task_counts = {
        person.id: count_tasks_for_person(db, camp, person)
        for person in people
    }

    person_missing_profile_fields = {
        person.id: get_missing_profile_fields(person)
        for person in people
    }

    return templates.TemplateResponse(
        "people/list.html",
        {
            "request": request,
            "camp": camp,
            "people": people,
            "participating_groups": participating_groups,
            "participating_group_lookup": participating_group_lookup,
            "sections": sections,
            "sections_by_group": sections_by_group,
            "sections_without_group": sections_without_group,
            "people_by_section": people_by_section,
            "people_without_section": people_without_section,
            "person_task_counts": person_task_counts,
            "person_missing_profile_fields": person_missing_profile_fields,
            "section_lookup": section_lookup,
        },
    )


@app.post("/camps/{camp_id}/people/bulk-edit")
async def bulk_edit_people(
    request: Request,
    camp_id: int,
    selected_person_id: list[int] | None = Form(None),
    bulk_action: str = Form(""),
    bulk_home_section_id: str = Form(""),
    bulk_person_type: str = Form(""),
    bulk_attendance_status: str = Form(""),
    bulk_provisional_status: str = Form(""),
    confirm_delete_people: str | None = Form(None),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    person_ids = selected_person_id or []

    if not person_ids:
        return RedirectResponse(
            url=f"/camps/{camp.id}/people?bulk_error=no_selection#bulk-actions",
            status_code=303,
        )

    people = (
        db.query(Person)
        .filter(Person.camp_id == camp.id, Person.id.in_(person_ids))
        .all()
    )

    if not people:
        return RedirectResponse(
            url=f"/camps/{camp.id}/people?bulk_error=no_valid_people#bulk-actions",
            status_code=303,
        )

    updated = 0

    if bulk_action == "section":
        selected_section_id = int(bulk_home_section_id) if bulk_home_section_id else None

        if selected_section_id is not None:
            section = (
                db.query(Section)
                .filter(Section.id == selected_section_id, Section.camp_id == camp.id)
                .first()
            )
            if section is None:
                return RedirectResponse(
                    url=f"/camps/{camp.id}/people?bulk_error=invalid_section#bulk-actions",
                    status_code=303,
                )

        for person in people:
            person.home_section_id = selected_section_id
            updated += 1

    elif bulk_action == "person_type":
        if bulk_person_type not in PERSON_TYPES:
            return RedirectResponse(
                url=f"/camps/{camp.id}/people?bulk_error=invalid_person_type#bulk-actions",
                status_code=303,
            )

        for person in people:
            person.person_type = bulk_person_type
            updated += 1

    elif bulk_action == "attendance_status":
        allowed_statuses = ["", "Invited", "Attending", "Not Attending", "Maybe", "Unknown"]

        if bulk_attendance_status not in allowed_statuses:
            return RedirectResponse(
                url=f"/camps/{camp.id}/people?bulk_error=invalid_attendance_status#bulk-actions",
                status_code=303,
            )

        for person in people:
            person.attendance_status = bulk_attendance_status or None
            updated += 1

    elif bulk_action == "provisional_status":
        if bulk_provisional_status not in ["provisional", "confirmed"]:
            return RedirectResponse(
                url=f"/camps/{camp.id}/people?bulk_error=invalid_provisional_status#bulk-actions",
                status_code=303,
            )

        for person in people:
            person.is_provisional = bulk_provisional_status == "provisional"
            updated += 1

    elif bulk_action == "delete":
        if confirm_delete_people != "yes":
            return RedirectResponse(
                url=f"/camps/{camp.id}/people?bulk_error=delete_not_confirmed#bulk-actions",
                status_code=303,
            )

        real_person_ids = [person.id for person in people]

        affected_task_ids = {
            row.task_id
            for row in db.query(TaskAssignment)
            .filter(
                TaskAssignment.camp_id == camp.id,
                TaskAssignment.assigned_person_id.in_(real_person_ids),
            )
            .all()
        }

        db.query(TeamMembership).filter(
            TeamMembership.camp_id == camp.id,
            TeamMembership.person_id.in_(real_person_ids),
        ).delete(synchronize_session=False)

        db.query(TaskAssignment).filter(
            TaskAssignment.camp_id == camp.id,
            TaskAssignment.assigned_person_id.in_(real_person_ids),
        ).delete(synchronize_session=False)

        db.query(Activity).filter(
            Activity.camp_id == camp.id,
            Activity.activity_lead_id.in_(real_person_ids),
        ).update(
            {Activity.activity_lead_id: None},
            synchronize_session=False,
        )

        db.query(ProgrammeSession).filter(
            ProgrammeSession.camp_id == camp.id,
            ProgrammeSession.lead_person_id.in_(real_person_ids),
        ).update(
            {ProgrammeSession.lead_person_id: None},
            synchronize_session=False,
        )

        deleted_count = len(people)

        for person in people:
            db.delete(person)

        db.flush()

        for task in db.query(Task).filter(Task.camp_id == camp.id, Task.id.in_(affected_task_ids)).all():
            update_task_status_from_assignments(db, task)

        db.commit()

        return RedirectResponse(
            url=f"/camps/{camp.id}/people?bulk_deleted={deleted_count}#bulk-actions",
            status_code=303,
        )

    else:
        return RedirectResponse(
            url=f"/camps/{camp.id}/people?bulk_error=invalid_action#bulk-actions",
            status_code=303,
        )

    db.commit()

    return RedirectResponse(
        url=f"/camps/{camp.id}/people?bulk_updated={updated}#bulk-actions",
        status_code=303,
    )



@app.post("/camps/{camp_id}/people/bulk-delete")
async def bulk_delete_people(
    request: Request,
    camp_id: int,
    selected_person_id: list[int] | None = Form(None),
    confirm_delete_people: str | None = Form(None),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    person_ids = selected_person_id or []

    if not person_ids:
        return RedirectResponse(
            url=f"/camps/{camp.id}/people?bulk_error=no_selection#bulk-actions",
            status_code=303,
        )

    if confirm_delete_people != "yes":
        return RedirectResponse(
            url=f"/camps/{camp.id}/people?bulk_error=delete_not_confirmed#bulk-actions",
            status_code=303,
        )

    people = (
        db.query(Person)
        .filter(Person.camp_id == camp.id, Person.id.in_(person_ids))
        .all()
    )

    if not people:
        return RedirectResponse(
            url=f"/camps/{camp.id}/people?bulk_error=no_valid_people#bulk-actions",
            status_code=303,
        )

    real_person_ids = [person.id for person in people]

    affected_task_ids = {
        row.task_id
        for row in db.query(TaskAssignment)
        .filter(
            TaskAssignment.camp_id == camp.id,
            TaskAssignment.assigned_person_id.in_(real_person_ids),
        )
        .all()
    }

    db.query(TeamMembership).filter(
        TeamMembership.camp_id == camp.id,
        TeamMembership.person_id.in_(real_person_ids),
    ).delete(synchronize_session=False)

    db.query(TaskAssignment).filter(
        TaskAssignment.camp_id == camp.id,
        TaskAssignment.assigned_person_id.in_(real_person_ids),
    ).delete(synchronize_session=False)

    deleted_count = len(people)

    for person in people:
        db.delete(person)

    db.flush()

    for task in db.query(Task).filter(Task.camp_id == camp.id, Task.id.in_(affected_task_ids)).all():
        update_task_status_from_assignments(db, task)

    db.commit()

    return RedirectResponse(
        url=f"/camps/{camp.id}/people?bulk_deleted={deleted_count}#bulk-actions",
        status_code=303,
    )


@app.get("/camps/{camp_id}/people/osm-member-import", response_class=HTMLResponse)
async def osm_member_import_form(
    request: Request,
    camp_id: int,
    section_id: int | None = None,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    sections = get_active_sections(db, camp)

    selected_section = None
    if section_id is not None:
        selected_section = (
            db.query(Section)
            .filter(Section.id == section_id, Section.camp_id == camp.id)
            .first()
        )

    return templates.TemplateResponse(
        "people/osm_member_import.html",
        {
            "request": request,
            "camp": camp,
            "sections": sections,
            "participating_group_lookup": get_participating_group_lookup(db, camp),
            "selected_section": selected_section,
            "selected_section_id": selected_section.id if selected_section else None,
            "error": "Selected section was not found." if section_id is not None and selected_section is None else None,
        },
    )


@app.post("/camps/{camp_id}/people/osm-member-import", response_class=HTMLResponse)
async def preview_osm_member_import(
    request: Request,
    camp_id: int,
    section_id: int = Form(...),
    default_person_type: str = Form("Young Person"),
    osm_file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    sections = get_active_sections(db, camp)

    section = (
        db.query(Section)
        .filter(Section.id == section_id, Section.camp_id == camp.id)
        .first()
    )

    if section is None:
        return templates.TemplateResponse(
            "people/osm_member_import.html",
            {
                "request": request,
                "camp": camp,
                "sections": sections,
                "participating_group_lookup": get_participating_group_lookup(db, camp),
                "error": "Please choose a valid section.",
            },
            status_code=400,
        )

    filename = osm_file.filename or ""

    if not filename.lower().endswith(".xlsx"):
        return templates.TemplateResponse(
            "people/osm_member_import.html",
            {
                "request": request,
                "camp": camp,
                "sections": sections,
                "participating_group_lookup": get_participating_group_lookup(db, camp),
                "error": "Please upload an OSM member export in .xlsx format.",
            },
            status_code=400,
        )

    try:
        candidates = parse_osm_member_export(await osm_file.read())
    except Exception as exc:
        return templates.TemplateResponse(
            "people/osm_member_import.html",
            {
                "request": request,
                "camp": camp,
                "sections": sections,
                "participating_group_lookup": get_participating_group_lookup(db, camp),
                "error": f"Could not read this OSM member export: {exc}",
            },
            status_code=400,
        )

    people_in_section = (
        db.query(Person)
        .filter(Person.camp_id == camp.id, Person.home_section_id == section.id)
        .order_by(Person.last_name, Person.first_name)
        .all()
    )

    lookup = {}
    for person in people_in_section:
        key = (
            normalise_person_match_value(person.first_name),
            normalise_person_match_value(person.last_name),
        )
        lookup.setdefault(key, []).append(person)

    people_elsewhere_in_camp = (
        db.query(Person)
        .filter(
            Person.camp_id == camp.id,
            (Person.home_section_id != section.id) | (Person.home_section_id.is_(None)),
        )
        .order_by(Person.last_name, Person.first_name)
        .all()
    )

    elsewhere_lookup = {}
    for person in people_elsewhere_in_camp:
        key = (
            normalise_person_match_value(person.first_name),
            normalise_person_match_value(person.last_name),
        )
        elsewhere_lookup.setdefault(key, []).append(person)

    section_lookup = {section.id: section for section in sections}
    participating_group_lookup_for_locations = get_participating_group_lookup(db, camp)

    preview_rows = []
    matched_count = 0
    unmatched_count = 0

    for candidate in candidates:
        key = (
            normalise_person_match_value(candidate.get("first_name")),
            normalise_person_match_value(candidate.get("last_name")),
        )

        matches = lookup.get(key, [])
        elsewhere_matches = elsewhere_lookup.get(key, [])
        matched_person = matches[0] if len(matches) == 1 else None
        name_exists_elsewhere = bool(elsewhere_matches)

        if matched_person:
            matched_count += 1
            warning = ""
        elif len(matches) > 1:
            unmatched_count += 1
            warning = "Multiple matches found in this section"
        elif name_exists_elsewhere:
            unmatched_count += 1
            warning = "Same name exists elsewhere in this camp, but not in this section"
        else:
            unmatched_count += 1
            warning = "No match found"

        elsewhere_match_summaries = build_osm_duplicate_location_summaries(
            camp,
            elsewhere_matches,
            section_lookup,
            participating_group_lookup_for_locations,
        )

        apply_payload = dict(candidate)
        apply_payload["matched_person_id"] = matched_person.id if matched_person else None
        apply_payload["suggested_person_type"] = candidate.get("suggested_person_type") or default_person_type
        apply_payload["name_exists_elsewhere"] = name_exists_elsewhere

        preview_rows.append(
            {
                **candidate,
                "matched_person": matched_person,
                "warning": warning,
                "elsewhere_matches": elsewhere_match_summaries,
                "payload": json.dumps(apply_payload),
            }
        )

    return templates.TemplateResponse(
        "people/osm_member_import.html",
        {
            "request": request,
            "camp": camp,
            "sections": sections,
            "participating_group_lookup": get_participating_group_lookup(db, camp),
            "selected_section": section,
            "selected_section_id": section.id,
            "default_person_type": default_person_type,
            "filename": filename,
            "preview_rows": preview_rows,
            "matched_count": matched_count,
            "unmatched_count": unmatched_count,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/people/osm-member-import/apply")
async def apply_osm_member_import(
    request: Request,
    camp_id: int,
    section_id: int = Form(...),
    default_person_type: str = Form("Young Person"),
    overwrite_existing: str | None = Form(None),
    replace_provisional: str | None = Form(None),
    create_missing: str | None = Form(None),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    section = (
        db.query(Section)
        .filter(Section.id == section_id, Section.camp_id == camp.id)
        .first()
    )

    if section is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Section not found."},
            status_code=404,
        )

    form = await request.form()
    selected_payloads = form.getlist("selected_payload")

    should_overwrite = overwrite_existing == "yes"

    updated = 0
    replaced = 0
    created = 0
    skipped = 0

    for payload_text in selected_payloads:
        data = json.loads(payload_text)

        person = None
        matched_person_id = data.get("matched_person_id")

        if matched_person_id:
            person = (
                db.query(Person)
                .filter(
                    Person.id == int(matched_person_id),
                    Person.camp_id == camp.id,
                    Person.home_section_id == section.id,
                )
                .first()
            )

        if person is not None:
            apply_osm_member_candidate_to_person(
                person,
                data,
                overwrite_existing=should_overwrite,
                update_identity=False,
            )
            updated += 1
            continue

        if data.get("name_exists_elsewhere"):
            skipped += 1
            continue

        if replace_provisional == "yes":
            person = (
                db.query(Person)
                .filter(
                    Person.camp_id == camp.id,
                    Person.home_section_id == section.id,
                    Person.is_provisional == True,
                )
                .order_by(Person.id)
                .first()
            )

            if person is not None:
                person.person_type = data.get("suggested_person_type") or default_person_type
                person.attendance_status = person.attendance_status or "Expected"

                apply_osm_member_candidate_to_person(
                    person,
                    data,
                    overwrite_existing=True,
                    update_identity=True,
                )

                replaced += 1
                continue

        if create_missing == "yes":
            first_name = clean_osm_cell(data.get("first_name")) or "Unknown"
            last_name = clean_osm_cell(data.get("last_name")) or "Unknown"

            person = Person(
                camp_id=camp.id,
                first_name=first_name,
                last_name=last_name,
                person_type=data.get("suggested_person_type") or default_person_type,
                home_section_id=section.id,
                is_provisional=False,
                attendance_status="Unknown",
                information_source="OSM Member File",
            )

            db.add(person)
            db.flush()

            apply_osm_member_candidate_to_person(
                person,
                data,
                overwrite_existing=True,
                update_identity=False,
            )

            created += 1
            continue

        skipped += 1

    db.commit()

    return RedirectResponse(
        url=f"/camps/{camp.id}/people?member_import_updated={updated}&member_import_replaced={replaced}&member_import_created={created}&member_import_skipped={skipped}",
        status_code=303,
    )


@app.get("/camps/{camp_id}/people/osm-attendance-update", response_class=HTMLResponse)
async def osm_attendance_update_form(
    request: Request,
    camp_id: int,
    section_id: int | None = None,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    sections = get_active_sections(db, camp)

    selected_section = None
    if section_id is not None:
        selected_section = (
            db.query(Section)
            .filter(Section.id == section_id, Section.camp_id == camp.id)
            .first()
        )

    return templates.TemplateResponse(
        "people/osm_attendance_update.html",
        {
            "request": request,
            "camp": camp,
            "sections": sections,
            "participating_group_lookup": get_participating_group_lookup(db, camp),
            "selected_section": selected_section,
            "selected_section_id": selected_section.id if selected_section else None,
            "error": "Selected section was not found." if section_id is not None and selected_section is None else None,
        },
    )


@app.post("/camps/{camp_id}/people/osm-attendance-update", response_class=HTMLResponse)
async def preview_osm_attendance_update(
    request: Request,
    camp_id: int,
    section_id: int = Form(...),
    default_person_type: str = Form("Young Person"),
    osm_file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    sections = get_active_sections(db, camp)

    section = (
        db.query(Section)
        .filter(Section.id == section_id, Section.camp_id == camp.id)
        .first()
    )

    if section is None:
        return templates.TemplateResponse(
            "people/osm_attendance_update.html",
            {
                "request": request,
                "camp": camp,
                "sections": sections,
                "participating_group_lookup": get_participating_group_lookup(db, camp),
                "error": "Please choose a valid section.",
            },
            status_code=400,
        )

    filename = osm_file.filename or ""

    if not filename.lower().endswith(".xlsx"):
        return templates.TemplateResponse(
            "people/osm_attendance_update.html",
            {
                "request": request,
                "camp": camp,
                "sections": sections,
                "participating_group_lookup": get_participating_group_lookup(db, camp),
                "error": "Please upload an OSM event export in .xlsx format.",
            },
            status_code=400,
        )

    try:
        rows = parse_osm_event_attendance_export(await osm_file.read())
    except Exception as exc:
        return templates.TemplateResponse(
            "people/osm_attendance_update.html",
            {
                "request": request,
                "camp": camp,
                "sections": sections,
                "participating_group_lookup": get_participating_group_lookup(db, camp),
                "error": f"Could not read this OSM event export: {exc}",
            },
            status_code=400,
        )

    people_in_section = (
        db.query(Person)
        .filter(Person.camp_id == camp.id, Person.home_section_id == section.id)
        .order_by(Person.last_name, Person.first_name)
        .all()
    )

    lookup = {}
    for person in people_in_section:
        key = (
            normalise_person_match_value(person.first_name),
            normalise_person_match_value(person.last_name),
        )
        lookup.setdefault(key, []).append(person)

    people_elsewhere_in_camp = (
        db.query(Person)
        .filter(
            Person.camp_id == camp.id,
            (Person.home_section_id != section.id) | (Person.home_section_id.is_(None)),
        )
        .order_by(Person.last_name, Person.first_name)
        .all()
    )

    elsewhere_lookup = {}
    for person in people_elsewhere_in_camp:
        key = (
            normalise_person_match_value(person.first_name),
            normalise_person_match_value(person.last_name),
        )
        elsewhere_lookup.setdefault(key, []).append(person)

    section_lookup = {section.id: section for section in sections}
    participating_group_lookup_for_locations = get_participating_group_lookup(db, camp)

    preview_rows = []
    matched_count = 0
    unmatched_count = 0

    for row in rows:
        key = (
            normalise_person_match_value(row["first_name"]),
            normalise_person_match_value(row["last_name"]),
        )

        matches = lookup.get(key, [])
        elsewhere_matches = elsewhere_lookup.get(key, [])
        matched_person = matches[0] if len(matches) == 1 else None
        name_exists_elsewhere = bool(elsewhere_matches)

        if matched_person:
            matched_count += 1
            warning = ""
        elif len(matches) > 1:
            unmatched_count += 1
            warning = "Multiple matches found in this section"
        elif name_exists_elsewhere:
            unmatched_count += 1
            warning = "Same name exists elsewhere in this camp, but not in this section"
        else:
            unmatched_count += 1
            warning = "No match found"

        elsewhere_match_summaries = build_osm_duplicate_location_summaries(
            camp,
            elsewhere_matches,
            section_lookup,
            participating_group_lookup_for_locations,
        )

        payload = {
            "first_name": row["first_name"],
            "last_name": row["last_name"],
            "attendance_status": row["attendance_status"],
            "attending_raw": row["attending_raw"],
            "matched_person_id": matched_person.id if matched_person else None,
            "name_exists_elsewhere": name_exists_elsewhere,
        }

        preview_rows.append(
            {
                **row,
                "matched_person": matched_person,
                "warning": warning,
                "elsewhere_matches": elsewhere_match_summaries,
                "payload": json.dumps(payload),
            }
        )

    return templates.TemplateResponse(
        "people/osm_attendance_update.html",
        {
            "request": request,
            "camp": camp,
            "sections": sections,
            "participating_group_lookup": get_participating_group_lookup(db, camp),
            "selected_section": section,
            "selected_section_id": section.id,
            "default_person_type": default_person_type,
            "filename": filename,
            "preview_rows": preview_rows,
            "matched_count": matched_count,
            "unmatched_count": unmatched_count,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/people/osm-attendance-update/apply")
async def apply_osm_attendance_update(
    request: Request,
    camp_id: int,
    section_id: int = Form(...),
    default_person_type: str = Form("Young Person"),
    create_missing: str | None = Form(None),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    section = (
        db.query(Section)
        .filter(Section.id == section_id, Section.camp_id == camp.id)
        .first()
    )

    if section is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Section not found."},
            status_code=404,
        )

    form = await request.form()
    selected_payloads = form.getlist("selected_payload")

    for payload_text in selected_payloads:
        data = json.loads(payload_text)
        matched_person_id = data.get("matched_person_id")
        person = None

        if matched_person_id:
            person = (
                db.query(Person)
                .filter(Person.id == int(matched_person_id), Person.camp_id == camp.id)
                .first()
            )

        if person:
            person.attendance_status = data.get("attendance_status") or "No response"
            person.information_source = "OSM Event Export"
            continue

        if data.get("name_exists_elsewhere"):
            continue

        if create_missing == "yes":
            first_name = (data.get("first_name") or "").strip()
            last_name = (data.get("last_name") or "").strip()

            if first_name or last_name:
                db.add(
                    Person(
                        camp_id=camp.id,
                        first_name=first_name or "Unknown",
                        last_name=last_name or "Unknown",
                        person_type=default_person_type,
                        home_section_id=section.id,
                        is_provisional=False,
                        attendance_status=data.get("attendance_status") or "No response",
                        information_source="OSM Event Export",
                    )
                )

    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/people", status_code=303)


@app.get("/camps/{camp_id}/people/provisional", response_class=HTMLResponse)
async def provisional_people_form(request: Request, camp_id: int, db: Session = Depends(get_db)):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    sections = get_active_sections(db, camp)

    return templates.TemplateResponse(
        "people/provisional.html",
        {
            "request": request,
            "camp": camp,
            "sections": sections,
            "participating_group_lookup": get_participating_group_lookup(db, camp),
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/people/provisional")
async def create_provisional_people(
    request: Request,
    camp_id: int,
    section_id: int = Form(...),
    count: int = Form(...),
    prefix: str = Form("YP"),
    person_type: str = Form("Young Person"),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    sections = get_active_sections(db, camp)

    section = (
        db.query(Section)
        .filter(Section.id == section_id, Section.camp_id == camp.id)
        .first()
    )

    if section is None:
        return templates.TemplateResponse(
            "people/provisional.html",
            {
                "request": request,
                "camp": camp,
                "sections": sections,
                "participating_group_lookup": get_participating_group_lookup(db, camp),
                "error": "Please choose a valid section.",
            },
            status_code=400,
        )

    if count < 1 or count > 100:
        return templates.TemplateResponse(
            "people/provisional.html",
            {
                "request": request,
                "camp": camp,
                "sections": sections,
                "error": "Please choose between 1 and 100 provisional people.",
            },
            status_code=400,
        )

    clean_prefix = (prefix or "YP").strip().upper()

    existing_people = (
        db.query(Person)
        .filter(Person.camp_id == camp.id, Person.home_section_id == section.id)
        .all()
    )

    used_numbers = set()
    for person in existing_people:
        if person.first_name == section.name and person.last_name.upper().startswith(clean_prefix):
            suffix = person.last_name[len(clean_prefix):]
            if suffix.isdigit():
                used_numbers.add(int(suffix))

    created = 0
    next_number = 1

    while created < count:
        while next_number in used_numbers:
            next_number += 1

        label = f"{clean_prefix}{next_number:02d}"

        db.add(
            Person(
                camp_id=camp.id,
                first_name=section.name,
                last_name=label,
                person_type=person_type,
                home_section_id=section.id,
                is_provisional=True,
                attendance_status="Provisional",
                information_source="Provisional",
                role_notes="Provisional attendee placeholder. Replace with real attendee details when known.",
            )
        )

        used_numbers.add(next_number)
        created += 1
        next_number += 1

    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/people", status_code=303)


@app.get("/camps/{camp_id}/people/new", response_class=HTMLResponse)
async def new_person_form(request: Request, camp_id: int, db: Session = Depends(get_db)):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Camp not found.",
            },
            status_code=404,
        )

    sections = get_active_sections(db, camp)

    return templates.TemplateResponse(
        "people/new.html",
        {
            "request": request,
            "camp": camp,
            "sections": sections,
            "participating_group_lookup": get_participating_group_lookup(db, camp),
            "duplicate_people": build_manual_duplicate_people_data(db, camp),
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/people/new")
async def create_person(
    request: Request,
    camp_id: int,
    first_name: str = Form(...),
    last_name: str = Form(...),
    person_type: str = Form(...),
    home_section_id: str = Form(""),
    attendance_status: str = Form("Expected"),
    information_source: str = Form("Manual Entry"),
    is_provisional: str | None = Form(None),
    email: str = Form(""),
    phone: str = Form(""),
    role_notes: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Camp not found.",
            },
            status_code=404,
        )

    if not first_name.strip() or not last_name.strip():
        return templates.TemplateResponse(
            "people/new.html",
            {
                "request": request,
                "camp": camp,
                "sections": get_active_sections(db, camp),
                "participating_group_lookup": get_participating_group_lookup(db, camp),
                "duplicate_people": build_manual_duplicate_people_data(db, camp),
                "form_data": {
                    "first_name": first_name,
                    "last_name": last_name,
                    "person_type": person_type,
                    "home_section_id": home_section_id,
                    "email": email,
                    "phone": phone,
                    "role_notes": role_notes,
                },
                "error": "First name and last name are required.",
            },
            status_code=400,
        )

    selected_home_section_id = int(home_section_id) if home_section_id else None

    person = Person(
        camp_id=camp.id,
        first_name=first_name.strip(),
        last_name=last_name.strip(),
        person_type=person_type,
        home_section_id=selected_home_section_id,
        is_provisional=False,
        attendance_status="Expected",
        information_source="Manual Entry",
        email=email.strip() or None,
        phone=phone.strip() or None,
        role_notes=role_notes.strip() or None,
    )

    db.add(person)
    db.commit()
    db.refresh(person)

    return RedirectResponse(url=f"/camps/{camp.id}/people/{person.id}", status_code=303)




@app.get("/camps/{camp_id}/people/{person_id}/osm-update", response_class=HTMLResponse)
async def osm_update_person_form(
    request: Request,
    camp_id: int,
    person_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    person = (
        db.query(Person)
        .filter(Person.id == person_id, Person.camp_id == camp.id)
        .first()
    )

    if person is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Person not found."},
            status_code=404,
        )

    home_section = None
    if person.home_section_id:
        home_section = (
            db.query(Section)
            .filter(Section.id == person.home_section_id, Section.camp_id == camp.id)
            .first()
        )

    return templates.TemplateResponse(
        "people/osm_update.html",
        {
            "request": request,
            "camp": camp,
            "person": person,
            "home_section": home_section,
        },
    )


@app.post("/camps/{camp_id}/people/{person_id}/osm-update", response_class=HTMLResponse)
async def preview_osm_update_person(
    request: Request,
    camp_id: int,
    person_id: int,
    osm_file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    person = (
        db.query(Person)
        .filter(Person.id == person_id, Person.camp_id == camp.id)
        .first()
    )

    if person is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Person not found."},
            status_code=404,
        )

    home_section = None
    if person.home_section_id:
        home_section = (
            db.query(Section)
            .filter(Section.id == person.home_section_id, Section.camp_id == camp.id)
            .first()
        )

    filename = osm_file.filename or ""

    if not filename.lower().endswith(".xlsx"):
        return templates.TemplateResponse(
            "people/osm_update.html",
            {
                "request": request,
                "camp": camp,
                "person": person,
                "home_section": home_section,
                "error": "Please upload an OSM Excel export in .xlsx format.",
                "candidates": [],
            },
            status_code=400,
        )

    file_bytes = await osm_file.read()

    try:
        candidates = parse_osm_member_export(file_bytes)
    except Exception as exc:
        return templates.TemplateResponse(
            "people/osm_update.html",
            {
                "request": request,
                "camp": camp,
                "person": person,
                "home_section": home_section,
                "error": f"Could not read this OSM export: {exc}",
                "candidates": [],
            },
            status_code=400,
        )

    return templates.TemplateResponse(
        "people/osm_update.html",
        {
            "request": request,
            "camp": camp,
            "person": person,
            "home_section": home_section,
            "error": None,
            "candidates": candidates,
            "filename": filename,
        },
    )


@app.post("/camps/{camp_id}/people/{person_id}/osm-update/apply")
async def apply_osm_update_person(
    request: Request,
    camp_id: int,
    person_id: int,
    selected_payload: str = Form(...),
    update_identity: str | None = Form(None),
    overwrite_existing: str | None = Form(None),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    person = (
        db.query(Person)
        .filter(Person.id == person_id, Person.camp_id == camp.id)
        .first()
    )

    if person is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Person not found."},
            status_code=404,
        )

    data = json.loads(selected_payload)
    should_overwrite = overwrite_existing == "yes"

    def apply_text_field(attribute_name):
        incoming = clean_osm_cell(data.get(attribute_name))
        existing = clean_osm_cell(getattr(person, attribute_name, ""))

        if incoming and (should_overwrite or not existing):
            setattr(person, attribute_name, incoming)

    if update_identity == "yes":
        if clean_osm_cell(data.get("first_name")):
            person.first_name = clean_osm_cell(data.get("first_name"))

        if clean_osm_cell(data.get("last_name")):
            person.last_name = clean_osm_cell(data.get("last_name"))

        person.is_provisional = False
        person.attendance_status = "Expected"

        if person.role_notes and "Provisional attendee placeholder" in person.role_notes:
            person.role_notes = None

    for field_name in [
        "email",
        "phone",
        "primary_contact_name",
        "primary_contact_relationship",
        "primary_contact_phone",
        "primary_contact_email",
        "emergency_contact_name",
        "emergency_contact_relationship",
        "emergency_contact_phone",
        "emergency_contact_email",
        "allergies",
        "allergy_action",
        "medication",
        "medical_notes",
        "dietary_requirements",
    ]:
        apply_text_field(field_name)

    person.information_source = "OSM Member File"

    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/people/{person.id}", status_code=303)


@app.get("/camps/{camp_id}/people/{person_id}/replace-provisional", response_class=HTMLResponse)
async def replace_provisional_person_form(
    request: Request,
    camp_id: int,
    person_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    person = (
        db.query(Person)
        .filter(Person.id == person_id, Person.camp_id == camp.id)
        .first()
    )

    if person is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Person not found."},
            status_code=404,
        )

    sections = get_active_sections(db, camp)

    return templates.TemplateResponse(
        "people/replace_provisional.html",
        {
            "request": request,
            "camp": camp,
            "person": person,
            "sections": sections,
            "participating_group_lookup": get_participating_group_lookup(db, camp),
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/people/{person_id}/replace-provisional")
async def replace_provisional_person(
    request: Request,
    camp_id: int,
    person_id: int,
    first_name: str = Form(...),
    last_name: str = Form(...),
    home_section_id: str = Form(""),
    email: str = Form(""),
    phone: str = Form(""),
    information_source: str = Form("Manual Entry"),
    role_notes: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    person = (
        db.query(Person)
        .filter(Person.id == person_id, Person.camp_id == camp.id)
        .first()
    )

    if person is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Person not found."},
            status_code=404,
        )

    if not first_name.strip() or not last_name.strip():
        return templates.TemplateResponse(
            "people/replace_provisional.html",
            {
                "request": request,
                "camp": camp,
                "person": person,
                "sections": get_active_sections(db, camp),
                "error": "First name and last name are required.",
            },
            status_code=400,
        )

    person.first_name = first_name.strip()
    person.last_name = last_name.strip()
    person.home_section_id = int(home_section_id) if home_section_id else None
    person.email = email.strip() or None
    person.phone = phone.strip() or None
    person.information_source = information_source.strip() or "Manual Entry"
    person.attendance_status = "Expected"
    person.is_provisional = False

    if role_notes.strip():
        person.role_notes = role_notes.strip()
    elif person.role_notes and "Provisional attendee placeholder" in person.role_notes:
        person.role_notes = None

    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/people/{person.id}", status_code=303)


@app.get("/camps/{camp_id}/people/{person_id}/edit", response_class=HTMLResponse)
async def edit_person_form(
    request: Request,
    camp_id: int,
    person_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Camp not found.",
            },
            status_code=404,
        )

    person = (
        db.query(Person)
        .filter(Person.id == person_id, Person.camp_id == camp.id)
        .first()
    )

    if person is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Person not found.",
            },
            status_code=404,
        )

    sections = get_active_sections(db, camp)

    return templates.TemplateResponse(
        "people/edit.html",
        {
            "request": request,
            "camp": camp,
            "person": person,
            "sections": sections,
            "participating_group_lookup": get_participating_group_lookup(db, camp),
            "duplicate_people": build_manual_duplicate_people_data(db, camp, exclude_person_id=person.id),
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/people/{person_id}/edit")
async def update_person(
    request: Request,
    camp_id: int,
    person_id: int,
    first_name: str = Form(...),
    last_name: str = Form(...),
    person_type: str = Form(...),
    home_section_id: str = Form(""),
    section_unit: str = Form(""),
    attendance_status: str = Form("Expected"),
    information_source: str = Form("Manual Entry"),
    is_provisional: str | None = Form(None),
    email: str = Form(""),
    phone: str = Form(""),
    primary_contact_name: str = Form(""),
    primary_contact_relationship: str = Form(""),
    primary_contact_phone: str = Form(""),
    primary_contact_email: str = Form(""),
    emergency_contact_name: str = Form(""),
    emergency_contact_relationship: str = Form(""),
    emergency_contact_phone: str = Form(""),
    emergency_contact_email: str = Form(""),
    allergies: str = Form(""),
    allergy_action: str = Form(""),
    medication: str = Form(""),
    medical_notes: str = Form(""),
    dietary_requirements: str = Form(""),
    role_notes: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Camp not found.",
            },
            status_code=404,
        )

    person = (
        db.query(Person)
        .filter(Person.id == person_id, Person.camp_id == camp.id)
        .first()
    )

    if person is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Person not found.",
            },
            status_code=404,
        )

    if not first_name.strip() or not last_name.strip():
        return templates.TemplateResponse(
            "people/edit.html",
            {
                "request": request,
                "camp": camp,
                "person": person,
                "sections": get_active_sections(db, camp),
                "participating_group_lookup": get_participating_group_lookup(db, camp),
                "duplicate_people": build_manual_duplicate_people_data(db, camp, exclude_person_id=person.id),
                "error": "First name and last name are required.",
            },
            status_code=400,
        )

    person.first_name = first_name.strip()
    person.last_name = last_name.strip()
    person.person_type = person_type
    person.home_section_id = int(home_section_id) if home_section_id else None
    person.section_unit = section_unit.strip() or None
    person.attendance_status = attendance_status.strip() or None
    person.information_source = information_source.strip() or None
    person.is_provisional = is_provisional == "yes"
    person.email = email.strip() or None
    person.phone = phone.strip() or None

    person.primary_contact_name = primary_contact_name.strip() or None
    person.primary_contact_relationship = primary_contact_relationship.strip() or None
    person.primary_contact_phone = primary_contact_phone.strip() or None
    person.primary_contact_email = primary_contact_email.strip() or None

    person.emergency_contact_name = emergency_contact_name.strip() or None
    person.emergency_contact_relationship = emergency_contact_relationship.strip() or None
    person.emergency_contact_phone = emergency_contact_phone.strip() or None
    person.emergency_contact_email = emergency_contact_email.strip() or None

    person.allergies = allergies.strip() or None
    person.allergy_action = allergy_action.strip() or None
    person.medication = medication.strip() or None
    person.medical_notes = medical_notes.strip() or None
    person.dietary_requirements = dietary_requirements.strip() or None

    person.role_notes = role_notes.strip() or None

    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/people/{person.id}", status_code=303)


@app.get("/camps/{camp_id}/people/{person_id}", response_class=HTMLResponse)
async def person_detail(
    request: Request,
    camp_id: int,
    person_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Camp not found.",
            },
            status_code=404,
        )

    person = (
        db.query(Person)
        .filter(Person.id == person_id, Person.camp_id == camp.id)
        .first()
    )

    if person is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Person not found.",
            },
            status_code=404,
        )

    home_section = None
    if person.home_section_id:
        home_section = (
            db.query(Section)
            .filter(Section.id == person.home_section_id, Section.camp_id == camp.id)
            .first()
        )

    team_rows = (
        db.query(TeamMembership, Team)
        .join(Team, Team.id == TeamMembership.team_id)
        .filter(
            TeamMembership.person_id == person.id,
            TeamMembership.camp_id == camp.id,
        )
        .order_by(Team.team_type, Team.name)
        .all()
    )

    team_ids = [team.id for membership, team in team_rows]

    direct_task_rows = (
        db.query(TaskAssignment, Task)
        .join(Task, Task.id == TaskAssignment.task_id)
        .filter(
            TaskAssignment.assigned_person_id == person.id,
            TaskAssignment.camp_id == camp.id,
        )
        .order_by(Task.status, Task.priority, Task.due_date, Task.title)
        .all()
    )

    team_task_rows = []

    if team_ids:
        team_task_rows = (
            db.query(TaskAssignment, Task, Team)
            .join(Task, Task.id == TaskAssignment.task_id)
            .join(Team, Team.id == TaskAssignment.assigned_team_id)
            .filter(
                TaskAssignment.assigned_team_id.in_(team_ids),
                TaskAssignment.camp_id == camp.id,
            )
            .order_by(Team.name, Task.status, Task.priority, Task.due_date, Task.title)
            .all()
        )

    missing_profile_fields = get_missing_profile_fields(person)

    return templates.TemplateResponse(
        "people/detail.html",
        {
            "request": request,
            "camp": camp,
            "person": person,
            "team_rows": team_rows,
            "home_section": home_section,
            "missing_profile_fields": missing_profile_fields,
            "direct_task_rows": direct_task_rows,
            "team_task_rows": team_task_rows,
        },
    )


@app.get("/teams", response_class=HTMLResponse)
async def teams_page(request: Request, db: Session = Depends(get_db)):
    camp = get_latest_camp(db)

    if camp is None:
        return templates.TemplateResponse("teams/no_camp.html", {"request": request})

    return RedirectResponse(url=f"/camps/{camp.id}/teams", status_code=303)


@app.get("/camps/{camp_id}/teams", response_class=HTMLResponse)
async def camp_team_list(request: Request, camp_id: int, db: Session = Depends(get_db)):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Camp not found.",
            },
            status_code=404,
        )

    teams = (
        db.query(Team)
        .filter(Team.camp_id == camp.id)
        .order_by(Team.team_type, Team.name)
        .all()
    )

    team_member_counts = {}
    for team in teams:
        team_member_counts[team.id] = (
            db.query(TeamMembership)
            .filter(TeamMembership.team_id == team.id, TeamMembership.camp_id == camp.id)
            .count()
        )

    team_task_counts = {
        team.id: count_tasks_for_team(db, camp, team)
        for team in teams
    }

    return templates.TemplateResponse(
        "teams/list.html",
        {
            "request": request,
            "camp": camp,
            "teams": teams,
            "team_member_counts": team_member_counts,
            "team_task_counts": team_task_counts,
        },
    )


@app.get("/camps/{camp_id}/teams/new", response_class=HTMLResponse)
async def new_team_form(request: Request, camp_id: int, db: Session = Depends(get_db)):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Camp not found.",
            },
            status_code=404,
        )

    return templates.TemplateResponse(
        "teams/new.html",
        {
            "request": request,
            "camp": camp,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/teams/new")
async def create_team(
    request: Request,
    camp_id: int,
    name: str = Form(...),
    team_type: str = Form(...),
    description: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Camp not found.",
            },
            status_code=404,
        )

    if not name.strip():
        return templates.TemplateResponse(
            "teams/new.html",
            {
                "request": request,
                "camp": camp,
                "error": "Team name is required.",
            },
            status_code=400,
        )

    team = Team(
        camp_id=camp.id,
        name=name.strip(),
        team_type=team_type,
        description=description.strip() or None,
    )

    db.add(team)
    db.commit()
    db.refresh(team)

    return RedirectResponse(url=f"/camps/{camp.id}/teams/{team.id}", status_code=303)




@app.get("/camps/{camp_id}/teams/{team_id}/edit", response_class=HTMLResponse)
async def edit_team_form(
    request: Request,
    camp_id: int,
    team_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Camp not found.",
            },
            status_code=404,
        )

    team = (
        db.query(Team)
        .filter(Team.id == team_id, Team.camp_id == camp.id)
        .first()
    )

    if team is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Team not found.",
            },
            status_code=404,
        )

    return templates.TemplateResponse(
        "teams/edit.html",
        {
            "request": request,
            "camp": camp,
            "team": team,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/teams/{team_id}/edit")
async def update_team(
    request: Request,
    camp_id: int,
    team_id: int,
    name: str = Form(...),
    team_type: str = Form(...),
    description: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Camp not found.",
            },
            status_code=404,
        )

    team = (
        db.query(Team)
        .filter(Team.id == team_id, Team.camp_id == camp.id)
        .first()
    )

    if team is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Team not found.",
            },
            status_code=404,
        )

    if not name.strip():
        return templates.TemplateResponse(
            "teams/edit.html",
            {
                "request": request,
                "camp": camp,
                "team": team,
                "error": "Team name is required.",
            },
            status_code=400,
        )

    team.name = name.strip()
    team.team_type = team_type
    team.description = description.strip() or None

    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/teams/{team.id}", status_code=303)


@app.get("/camps/{camp_id}/teams/{team_id}", response_class=HTMLResponse)
async def team_detail(
    request: Request,
    camp_id: int,
    team_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Camp not found.",
            },
            status_code=404,
        )

    team = (
        db.query(Team)
        .filter(Team.id == team_id, Team.camp_id == camp.id)
        .first()
    )

    if team is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Team not found.",
            },
            status_code=404,
        )

    memberships = (
        db.query(TeamMembership, Person)
        .join(Person, Person.id == TeamMembership.person_id)
        .filter(TeamMembership.team_id == team.id, TeamMembership.camp_id == camp.id)
        .order_by(Person.last_name, Person.first_name)
        .all()
    )

    task_rows = (
        db.query(TaskAssignment, Task)
        .join(Task, Task.id == TaskAssignment.task_id)
        .filter(
            TaskAssignment.assigned_team_id == team.id,
            TaskAssignment.camp_id == camp.id,
        )
        .order_by(Task.status, Task.priority, Task.due_date, Task.title)
        .all()
    )

    return templates.TemplateResponse(
        "teams/detail.html",
        {
            "request": request,
            "camp": camp,
            "team": team,
            "memberships": memberships,
            "task_rows": task_rows,
        },
    )


@app.get("/camps/{camp_id}/teams/{team_id}/members/new", response_class=HTMLResponse)
async def add_team_member_form(
    request: Request,
    camp_id: int,
    team_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Camp not found.",
            },
            status_code=404,
        )

    team = (
        db.query(Team)
        .filter(Team.id == team_id, Team.camp_id == camp.id)
        .first()
    )

    if team is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Team not found.",
            },
            status_code=404,
        )

    existing_person_ids = [
        row.person_id
        for row in db.query(TeamMembership)
        .filter(TeamMembership.team_id == team.id, TeamMembership.camp_id == camp.id)
        .all()
    ]

    people_query = db.query(Person).filter(Person.camp_id == camp.id)

    if existing_person_ids:
        people_query = people_query.filter(Person.id.notin_(existing_person_ids))

    available_people = people_query.order_by(Person.person_type, Person.last_name, Person.first_name).all()

    return templates.TemplateResponse(
        "teams/add_member.html",
        {
            "request": request,
            "camp": camp,
            "team": team,
            "available_people": available_people,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/teams/{team_id}/members/new")
async def add_team_member(
    request: Request,
    camp_id: int,
    team_id: int,
    person_id: int = Form(...),
    role_in_team: str = Form(""),
    notes: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Camp not found.",
            },
            status_code=404,
        )

    team = (
        db.query(Team)
        .filter(Team.id == team_id, Team.camp_id == camp.id)
        .first()
    )

    person = (
        db.query(Person)
        .filter(Person.id == person_id, Person.camp_id == camp.id)
        .first()
    )

    if team is None or person is None:
        return templates.TemplateResponse(
            "not_found.html",
            {
                "request": request,
                "message": "Team or person not found.",
            },
            status_code=404,
        )

    existing = (
        db.query(TeamMembership)
        .filter(
            TeamMembership.team_id == team.id,
            TeamMembership.person_id == person.id,
            TeamMembership.camp_id == camp.id,
        )
        .first()
    )

    if existing is None:
        membership = TeamMembership(
            camp_id=camp.id,
            team_id=team.id,
            person_id=person.id,
            role_in_team=role_in_team.strip() or None,
            notes=notes.strip() or None,
        )
        db.add(membership)
        db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/teams/{team.id}", status_code=303)




@app.get("/camps/{camp_id}/teams/{team_id}/members/{membership_id}/edit", response_class=HTMLResponse)
async def edit_team_member_form(
    request: Request,
    camp_id: int,
    team_id: int,
    membership_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    team = (
        db.query(Team)
        .filter(Team.id == team_id, Team.camp_id == camp.id)
        .first()
    )

    if team is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Team not found."},
            status_code=404,
        )

    membership = (
        db.query(TeamMembership)
        .filter(
            TeamMembership.id == membership_id,
            TeamMembership.team_id == team.id,
            TeamMembership.camp_id == camp.id,
        )
        .first()
    )

    if membership is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Team membership not found."},
            status_code=404,
        )

    person = (
        db.query(Person)
        .filter(Person.id == membership.person_id, Person.camp_id == camp.id)
        .first()
    )

    if person is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Person not found."},
            status_code=404,
        )

    return templates.TemplateResponse(
        "teams/edit_member.html",
        {
            "request": request,
            "camp": camp,
            "team": team,
            "membership": membership,
            "person": person,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/teams/{team_id}/members/{membership_id}/edit")
async def update_team_member(
    request: Request,
    camp_id: int,
    team_id: int,
    membership_id: int,
    role_in_team: str = Form(""),
    notes: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    team = (
        db.query(Team)
        .filter(Team.id == team_id, Team.camp_id == camp.id)
        .first()
    )

    if team is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Team not found."},
            status_code=404,
        )

    membership = (
        db.query(TeamMembership)
        .filter(
            TeamMembership.id == membership_id,
            TeamMembership.team_id == team.id,
            TeamMembership.camp_id == camp.id,
        )
        .first()
    )

    if membership is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Team membership not found."},
            status_code=404,
        )

    membership.role_in_team = role_in_team.strip() or None
    membership.notes = notes.strip() or None

    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/teams/{team.id}", status_code=303)


@app.post("/camps/{camp_id}/teams/{team_id}/members/{membership_id}/remove")
async def remove_team_member(
    camp_id: int,
    team_id: int,
    membership_id: int,
    db: Session = Depends(get_db),
):
    membership = (
        db.query(TeamMembership)
        .filter(
            TeamMembership.id == membership_id,
            TeamMembership.camp_id == camp_id,
            TeamMembership.team_id == team_id,
        )
        .first()
    )

    if membership is not None:
        db.delete(membership)
        db.commit()

    return RedirectResponse(url=f"/camps/{camp_id}/teams/{team_id}", status_code=303)



@app.get("/tasks", response_class=HTMLResponse)
async def tasks_page(request: Request, db: Session = Depends(get_db)):
    camp = get_latest_camp(db)

    if camp is None:
        return templates.TemplateResponse("tasks/no_camp.html", {"request": request})

    return RedirectResponse(url=f"/camps/{camp.id}/tasks", status_code=303)






@app.get("/camps/{camp_id}/task-categories", response_class=HTMLResponse)
async def task_category_list(request: Request, camp_id: int, db: Session = Depends(get_db)):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    ensure_default_task_categories(db, camp)
    apply_default_task_category_descriptions(db, camp)

    categories = (
        db.query(TaskCategory)
        .filter(TaskCategory.camp_id == camp.id)
        .order_by(TaskCategory.sort_order, TaskCategory.name)
        .all()
    )

    category_task_counts = {}
    for category in categories:
        category_task_counts[category.id] = (
            db.query(Task)
            .filter(Task.camp_id == camp.id, Task.category == category.name)
            .count()
        )

    return templates.TemplateResponse(
        "task_categories/list.html",
        {
            "request": request,
            "camp": camp,
            "categories": categories,
            "category_task_counts": category_task_counts,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/task-categories/new")
async def create_task_category(
    request: Request,
    camp_id: int,
    name: str = Form(...),
    description: str = Form(""),
    sort_order: int = Form(0),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    clean_name = name.strip()

    if clean_name:
        existing = (
            db.query(TaskCategory)
            .filter(TaskCategory.camp_id == camp.id, TaskCategory.name == clean_name)
            .first()
        )

        if existing is None:
            db.add(
                TaskCategory(
                    camp_id=camp.id,
                    name=clean_name,
                    description=description.strip() or None,
                    sort_order=sort_order,
                    is_active=True,
                )
            )
            db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/task-categories", status_code=303)


@app.get("/camps/{camp_id}/task-categories/{category_id}/edit", response_class=HTMLResponse)
async def edit_task_category_form(
    request: Request,
    camp_id: int,
    category_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    category = (
        db.query(TaskCategory)
        .filter(TaskCategory.id == category_id, TaskCategory.camp_id == camp.id)
        .first()
    )

    if category is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Task category not found."},
            status_code=404,
        )

    return templates.TemplateResponse(
        "task_categories/edit.html",
        {
            "request": request,
            "camp": camp,
            "category": category,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/task-categories/{category_id}/edit")
async def update_task_category(
    request: Request,
    camp_id: int,
    category_id: int,
    name: str = Form(...),
    description: str = Form(""),
    sort_order: int = Form(0),
    is_active: str = Form("off"),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    category = (
        db.query(TaskCategory)
        .filter(TaskCategory.id == category_id, TaskCategory.camp_id == camp.id)
        .first()
    )

    if category is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Task category not found."},
            status_code=404,
        )

    clean_name = name.strip()

    if not clean_name:
        return templates.TemplateResponse(
            "task_categories/edit.html",
            {
                "request": request,
                "camp": camp,
                "category": category,
                "error": "Category name is required.",
            },
            status_code=400,
        )

    old_name = category.name

    category.name = clean_name
    category.description = description.strip() or None
    category.sort_order = sort_order
    category.is_active = is_active == "on"

    if old_name != clean_name:
        tasks_using_category = (
            db.query(Task)
            .filter(Task.camp_id == camp.id, Task.category == old_name)
            .all()
        )

        for task in tasks_using_category:
            task.category = clean_name

    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/task-categories", status_code=303)


@app.post("/camps/{camp_id}/task-categories/{category_id}/delete")
async def delete_task_category(
    request: Request,
    camp_id: int,
    category_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    category = (
        db.query(TaskCategory)
        .filter(TaskCategory.id == category_id, TaskCategory.camp_id == camp.id)
        .first()
    )

    if category is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Task category not found."},
            status_code=404,
        )

    task_count = (
        db.query(Task)
        .filter(Task.camp_id == camp.id, Task.category == category.name)
        .count()
    )

    if task_count == 0:
        db.delete(category)
        db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/task-categories", status_code=303)


@app.get("/camps/{camp_id}/task-phases", response_class=HTMLResponse)
async def task_phase_list(request: Request, camp_id: int, db: Session = Depends(get_db)):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    ensure_default_task_phases(db, camp)
    apply_default_task_phase_descriptions(db, camp)

    phases = (
        db.query(TaskPhase)
        .filter(TaskPhase.camp_id == camp.id)
        .order_by(TaskPhase.sort_order, TaskPhase.name)
        .all()
    )

    phase_task_counts = {}
    for phase in phases:
        phase_task_counts[phase.id] = (
            db.query(Task)
            .filter(Task.camp_id == camp.id, Task.phase == phase.name)
            .count()
        )

    return templates.TemplateResponse(
        "task_phases/list.html",
        {
            "request": request,
            "camp": camp,
            "phases": phases,
            "phase_task_counts": phase_task_counts,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/task-phases/new")
async def create_task_phase(
    request: Request,
    camp_id: int,
    name: str = Form(...),
    description: str = Form(""),
    sort_order: int = Form(0),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    clean_name = name.strip()

    if not clean_name:
        return RedirectResponse(url=f"/camps/{camp.id}/task-phases", status_code=303)

    existing = (
        db.query(TaskPhase)
        .filter(TaskPhase.camp_id == camp.id, TaskPhase.name == clean_name)
        .first()
    )

    if existing is None:
        db.add(
            TaskPhase(
                camp_id=camp.id,
                name=clean_name,
                description=description.strip() or None,
                sort_order=sort_order,
                is_active=True,
            )
        )
        db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/task-phases", status_code=303)


@app.post("/camps/{camp_id}/task-phases/{phase_id}/delete")
async def delete_task_phase(
    request: Request,
    camp_id: int,
    phase_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    phase = (
        db.query(TaskPhase)
        .filter(TaskPhase.id == phase_id, TaskPhase.camp_id == camp.id)
        .first()
    )

    if phase is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Task phase not found."},
            status_code=404,
        )

    task_count = (
        db.query(Task)
        .filter(Task.camp_id == camp.id, Task.phase == phase.name)
        .count()
    )

    if task_count > 0:
        phases = (
            db.query(TaskPhase)
            .filter(TaskPhase.camp_id == camp.id)
            .order_by(TaskPhase.sort_order, TaskPhase.name)
            .all()
        )

        phase_task_counts = {}
        for item in phases:
            phase_task_counts[item.id] = (
                db.query(Task)
                .filter(Task.camp_id == camp.id, Task.phase == item.name)
                .count()
            )

        return templates.TemplateResponse(
            "task_phases/list.html",
            {
                "request": request,
                "camp": camp,
                "phases": phases,
                "phase_task_counts": phase_task_counts,
                "error": f"Cannot delete '{phase.name}' because {task_count} task(s) still use it. Move those tasks to another phase first.",
            },
            status_code=400,
        )

    db.delete(phase)
    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/task-phases", status_code=303)


@app.get("/camps/{camp_id}/task-phases/{phase_id}/edit", response_class=HTMLResponse)
async def edit_task_phase_form(
    request: Request,
    camp_id: int,
    phase_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    phase = (
        db.query(TaskPhase)
        .filter(TaskPhase.id == phase_id, TaskPhase.camp_id == camp.id)
        .first()
    )

    if phase is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Task phase not found."},
            status_code=404,
        )

    return templates.TemplateResponse(
        "task_phases/edit.html",
        {
            "request": request,
            "camp": camp,
            "phase": phase,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/task-phases/{phase_id}/edit")
async def update_task_phase(
    request: Request,
    camp_id: int,
    phase_id: int,
    name: str = Form(...),
    description: str = Form(""),
    sort_order: int = Form(0),
    is_active: str = Form("off"),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    phase = (
        db.query(TaskPhase)
        .filter(TaskPhase.id == phase_id, TaskPhase.camp_id == camp.id)
        .first()
    )

    if phase is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Task phase not found."},
            status_code=404,
        )

    clean_name = name.strip()

    if not clean_name:
        return templates.TemplateResponse(
            "task_phases/edit.html",
            {
                "request": request,
                "camp": camp,
                "phase": phase,
                "error": "Phase name is required.",
            },
            status_code=400,
        )

    old_name = phase.name

    phase.name = clean_name
    phase.description = description.strip() or None
    phase.sort_order = sort_order
    phase.is_active = is_active == "on"

    if old_name != clean_name:
        tasks_using_phase = (
            db.query(Task)
            .filter(Task.camp_id == camp.id, Task.phase == old_name)
            .all()
        )

        for task in tasks_using_phase:
            task.phase = clean_name

    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/task-phases", status_code=303)


@app.get("/camps/{camp_id}/tasks", response_class=HTMLResponse)
async def camp_task_list(
    request: Request,
    camp_id: int,
    view: str = "",
    status: str = "",
    priority: str = "",
    phase: str = "",
    category: str = "",
    assignee: str = "",
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    all_tasks = (
        db.query(Task)
        .filter(Task.camp_id == camp.id)
        .order_by(Task.status, Task.priority, Task.due_date, Task.title)
        .all()
    )

    people = (
        db.query(Person)
        .filter(Person.camp_id == camp.id)
        .order_by(Person.person_type, Person.last_name, Person.first_name)
        .all()
    )

    teams = (
        db.query(Team)
        .filter(Team.camp_id == camp.id)
        .order_by(Team.team_type, Team.name)
        .all()
    )

    assignment_counts = {}
    for task in all_tasks:
        assignment_counts[task.id] = (
            db.query(TaskAssignment)
            .filter(TaskAssignment.task_id == task.id, TaskAssignment.camp_id == camp.id)
            .count()
        )

    today = date.today()
    complete_statuses = {"Done"}

    unassigned_tasks = [
        task for task in all_tasks
        if assignment_counts.get(task.id, 0) == 0 or task.status == "Unassigned"
    ]

    blocked_tasks = [
        task for task in all_tasks
        if task.status == "Blocked"
    ]

    high_priority_tasks = [
        task for task in all_tasks
        if task.priority in {"High", "Urgent"} and task.status not in complete_statuses
    ]

    overdue_tasks = [
        task for task in all_tasks
        if task.due_date and task.due_date < today and task.status not in complete_statuses
    ]

    due_soon_tasks = [
        task for task in all_tasks
        if task.due_date
        and today <= task.due_date
        and (task.due_date - today).days <= 7
        and task.status not in complete_statuses
    ]

    tasks = list(all_tasks)
    active_parts = []

    if view == "unassigned":
        tasks = unassigned_tasks
        active_parts.append("Needs owner")
    elif view == "blocked":
        tasks = blocked_tasks
        active_parts.append("Blocked")
    elif view == "high-priority":
        tasks = high_priority_tasks
        active_parts.append("High / urgent")
    elif view == "overdue":
        tasks = overdue_tasks
        active_parts.append("Overdue")
    elif view == "due-soon":
        tasks = due_soon_tasks
        active_parts.append("Due soon")

    if assignee:
        task_ids_for_assignee = set()

        try:
            kind, raw_id = assignee.split(":", 1)
            target_id = int(raw_id)
        except ValueError:
            kind = ""
            target_id = 0

        if kind == "person":
            person = (
                db.query(Person)
                .filter(Person.id == target_id, Person.camp_id == camp.id)
                .first()
            )

            if person:
                active_parts.append(f"Person: {person.first_name} {person.last_name}")

                direct_task_ids = [
                    row.task_id
                    for row in db.query(TaskAssignment)
                    .filter(
                        TaskAssignment.camp_id == camp.id,
                        TaskAssignment.assigned_person_id == person.id,
                    )
                    .all()
                ]

                team_ids = [
                    row.team_id
                    for row in db.query(TeamMembership)
                    .filter(
                        TeamMembership.camp_id == camp.id,
                        TeamMembership.person_id == person.id,
                    )
                    .all()
                ]

                team_task_ids = []

                if team_ids:
                    team_task_ids = [
                        row.task_id
                        for row in db.query(TaskAssignment)
                        .filter(
                            TaskAssignment.camp_id == camp.id,
                            TaskAssignment.assigned_team_id.in_(team_ids),
                        )
                        .all()
                    ]

                task_ids_for_assignee = set(direct_task_ids + team_task_ids)

        elif kind == "team":
            team = (
                db.query(Team)
                .filter(Team.id == target_id, Team.camp_id == camp.id)
                .first()
            )

            if team:
                active_parts.append(f"Team: {team.name}")

                task_ids_for_assignee = {
                    row.task_id
                    for row in db.query(TaskAssignment)
                    .filter(
                        TaskAssignment.camp_id == camp.id,
                        TaskAssignment.assigned_team_id == team.id,
                    )
                    .all()
                }

        tasks = [task for task in tasks if task.id in task_ids_for_assignee]

    if status:
        tasks = [task for task in tasks if task.status == status]
        active_parts.append(f"Status: {status}")

    if priority:
        tasks = [task for task in tasks if task.priority == priority]
        active_parts.append(f"Priority: {priority}")

    if phase:
        tasks = [task for task in tasks if (task.phase or "") == phase]
        active_parts.append(f"Phase: {phase}")

    if category:
        tasks = [task for task in tasks if (task.category or "") == category]
        active_parts.append(f"Category: {category}")

    statuses = TASK_STATUSES
    priorities = ["Urgent", "High", "Normal", "Low"]
    phases = [phase.name for phase in get_active_task_phases(db, camp)]
    categories = [category.name for category in get_active_task_categories(db, camp)]

    active_label = " · ".join(active_parts) if active_parts else "All tasks"

    printable_url = None
    printable_label = None

    has_extra_filters = bool(status or priority)

    if view == "unassigned" and not has_extra_filters and not assignee and not phase and not category:
        printable_url = f"/camps/{camp.id}/task-sheets/unassigned"
        printable_label = "Print unassigned task sheet"

    elif assignee and not has_extra_filters and not view and not phase and not category:
        try:
            kind, raw_id = assignee.split(":", 1)
            target_id = int(raw_id)
        except ValueError:
            kind = ""
            target_id = 0

        if kind == "person":
            person = (
                db.query(Person)
                .filter(Person.id == target_id, Person.camp_id == camp.id)
                .first()
            )
            if person:
                printable_url = f"/camps/{camp.id}/task-sheets/person/{person.id}"
                printable_label = "Print person task sheet"

        elif kind == "team":
            team = (
                db.query(Team)
                .filter(Team.id == target_id, Team.camp_id == camp.id)
                .first()
            )
            if team:
                printable_url = f"/camps/{camp.id}/task-sheets/team/{team.id}"
                printable_label = "Print team task sheet"

    elif phase and not has_extra_filters and not view and not assignee and not category:
        phase_record = (
            db.query(TaskPhase)
            .filter(TaskPhase.camp_id == camp.id, TaskPhase.name == phase)
            .first()
        )
        if phase_record:
            printable_url = f"/camps/{camp.id}/task-sheets/phase/{phase_record.id}"
            printable_label = "Print phase task sheet"

    elif category and not has_extra_filters and not view and not assignee and not phase:
        category_record = (
            db.query(TaskCategory)
            .filter(TaskCategory.camp_id == camp.id, TaskCategory.name == category)
            .first()
        )
        if category_record:
            printable_url = f"/camps/{camp.id}/task-sheets/category/{category_record.id}"
            printable_label = "Print category task sheet"

    return templates.TemplateResponse(
        "tasks/list.html",
        {
            "request": request,
            "camp": camp,
            "tasks": tasks,
            "assignment_counts": assignment_counts,
            "active_view": view,
            "active_status": status,
            "active_priority": priority,
            "active_phase": phase,
            "active_category": category,
            "active_assignee": assignee,
            "active_label": active_label,
            "printable_url": printable_url,
            "printable_label": printable_label,
            "statuses": statuses,
            "priorities": priorities,
            "phases": phases,
            "categories": categories,
            "people": people,
            "teams": teams,
            "summary": {
                "total": len(all_tasks),
                "unassigned": len(unassigned_tasks),
                "blocked": len(blocked_tasks),
                "high_priority": len(high_priority_tasks),
                "overdue": len(overdue_tasks),
                "due_soon": len(due_soon_tasks),
            },
        },
    )


@app.get("/camps/{camp_id}/tasks/new", response_class=HTMLResponse)
async def new_task_form(request: Request, camp_id: int, db: Session = Depends(get_db)):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    phases = get_active_task_phases(db, camp)
    categories = get_active_task_categories(db, camp)

    return templates.TemplateResponse(
        "tasks/new.html",
        {
            "request": request,
            "camp": camp,
            "phases": phases,
            "categories": categories,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/tasks/new")
async def create_task(
    request: Request,
    camp_id: int,
    title: str = Form(...),
    description: str = Form(""),
    category: str = Form(""),
    phase: str = Form(""),
    priority: str = Form("Normal"),
    status: str = Form("To Do"),
    due_date: str = Form(""),
    notes: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    if not title.strip():
        phases = get_active_task_phases(db, camp)
        categories = get_active_task_categories(db, camp)

        return templates.TemplateResponse(
            "tasks/new.html",
            {
                "request": request,
                "camp": camp,
                "phases": phases,
                "error": "Task title is required.",
            },
            status_code=400,
        )

    task = Task(
        camp_id=camp.id,
        title=title.strip(),
        description=description.strip() or None,
        category=category.strip() or None,
        phase=phase.strip() or None,
        priority=priority,
        status=status,
        due_date=parse_optional_date(due_date),
        notes=notes.strip() or None,
    )

    db.add(task)
    db.commit()
    db.refresh(task)

    return RedirectResponse(url=f"/camps/{camp.id}/tasks/{task.id}", status_code=303)




@app.get("/camps/{camp_id}/tasks/{task_id}/edit", response_class=HTMLResponse)
async def edit_task_form(
    request: Request,
    camp_id: int,
    task_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    task = (
        db.query(Task)
        .filter(Task.id == task_id, Task.camp_id == camp.id)
        .first()
    )

    if task is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Task not found."},
            status_code=404,
        )

    phases = get_active_task_phases(db, camp)
    categories = get_active_task_categories(db, camp)

    return templates.TemplateResponse(
        "tasks/edit.html",
        {
            "request": request,
            "camp": camp,
            "task": task,
            "phases": phases,
            "categories": categories,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/tasks/{task_id}/edit")
async def update_task(
    request: Request,
    camp_id: int,
    task_id: int,
    title: str = Form(...),
    description: str = Form(""),
    category: str = Form(""),
    phase: str = Form(""),
    priority: str = Form("Normal"),
    status: str = Form("To Do"),
    due_date: str = Form(""),
    notes: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    task = (
        db.query(Task)
        .filter(Task.id == task_id, Task.camp_id == camp.id)
        .first()
    )

    if task is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Task not found."},
            status_code=404,
        )

    if not title.strip():
        phases = get_active_task_phases(db, camp)
        categories = get_active_task_categories(db, camp)

        return templates.TemplateResponse(
            "tasks/edit.html",
            {
                "request": request,
                "camp": camp,
                "task": task,
                "phases": phases,
                "error": "Task title is required.",
            },
            status_code=400,
        )

    task.title = title.strip()
    task.description = description.strip() or None
    task.category = category.strip() or None
    task.phase = phase.strip() or None
    task.priority = priority
    task.status = status
    task.due_date = parse_optional_date(due_date)
    task.notes = notes.strip() or None

    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/tasks/{task.id}", status_code=303)


@app.get("/camps/{camp_id}/tasks/{task_id}", response_class=HTMLResponse)
async def task_detail(
    request: Request,
    camp_id: int,
    task_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    task = (
        db.query(Task)
        .filter(Task.id == task_id, Task.camp_id == camp.id)
        .first()
    )

    if task is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Task not found."},
            status_code=404,
        )

    assignments = (
        db.query(TaskAssignment)
        .filter(TaskAssignment.task_id == task.id, TaskAssignment.camp_id == camp.id)
        .order_by(TaskAssignment.id)
        .all()
    )

    assignment_rows = []
    for assignment in assignments:
        assignee_label = "Unknown"
        assignee_type = ""

        if assignment.assigned_person_id:
            person = db.get(Person, assignment.assigned_person_id)
            if person:
                assignee_label = f"{person.first_name} {person.last_name}"
                assignee_type = person.person_type

        if assignment.assigned_team_id:
            team = db.get(Team, assignment.assigned_team_id)
            if team:
                assignee_label = team.name
                assignee_type = team.team_type

        assignment_rows.append(
            {
                "assignment": assignment,
                "assignee_label": assignee_label,
                "assignee_type": assignee_type,
            }
        )

    return templates.TemplateResponse(
        "tasks/detail.html",
        {
            "request": request,
            "camp": camp,
            "task": task,
            "assignment_rows": assignment_rows,
        },
    )


@app.get("/camps/{camp_id}/tasks/{task_id}/assignments/new", response_class=HTMLResponse)
async def new_task_assignment_form(
    request: Request,
    camp_id: int,
    task_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    task = (
        db.query(Task)
        .filter(Task.id == task_id, Task.camp_id == camp.id)
        .first()
    )

    if task is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Task not found."},
            status_code=404,
        )

    people, teams = get_available_task_assignees(db, camp, task)

    return templates.TemplateResponse(
        "tasks/add_assignment.html",
        {
            "request": request,
            "camp": camp,
            "task": task,
            "people": people,
            "teams": teams,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/tasks/{task_id}/assignments/new")
async def create_task_assignment(
    request: Request,
    camp_id: int,
    task_id: int,
    assignee: str = Form(...),
    status_override: str = Form(""),
    assignment_notes: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    task = (
        db.query(Task)
        .filter(Task.id == task_id, Task.camp_id == camp.id)
        .first()
    )

    if task is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Task not found."},
            status_code=404,
        )

    kind, raw_id = assignee.split(":", 1)
    target_id = int(raw_id)

    assigned_person_id = None
    assigned_team_id = None

    if kind == "person":
        person = (
            db.query(Person)
            .filter(Person.id == target_id, Person.camp_id == camp.id)
            .first()
        )
        if person is None:
            return templates.TemplateResponse(
                "not_found.html",
                {"request": request, "message": "Person not found."},
                status_code=404,
            )

        assigned_person_id = person.id

    elif kind == "team":
        team = (
            db.query(Team)
            .filter(Team.id == target_id, Team.camp_id == camp.id)
            .first()
        )
        if team is None:
            return templates.TemplateResponse(
                "not_found.html",
                {"request": request, "message": "Team not found."},
                status_code=404,
            )

        assigned_team_id = team.id

    else:
        people, teams = get_available_task_assignees(db, camp, task)
        return templates.TemplateResponse(
            "tasks/add_assignment.html",
            {
                "request": request,
                "camp": camp,
                "task": task,
                "people": people,
                "teams": teams,
                "error": "Choose a valid person or team.",
            },
            status_code=400,
        )

    if task_assignment_duplicate_exists(
        db,
        task,
        assigned_person_id=assigned_person_id,
        assigned_team_id=assigned_team_id,
    ):
        people, teams = get_available_task_assignees(db, camp, task)
        return templates.TemplateResponse(
            "tasks/add_assignment.html",
            {
                "request": request,
                "camp": camp,
                "task": task,
                "people": people,
                "teams": teams,
                "error": "This task is already assigned to that person or team.",
            },
            status_code=400,
        )

    assignment = TaskAssignment(
        camp_id=camp.id,
        task_id=task.id,
        assigned_person_id=assigned_person_id,
        assigned_team_id=assigned_team_id,
        status_override=status_override.strip() or None,
        assignment_notes=assignment_notes.strip() or None,
    )

    db.add(assignment)
    db.flush()
    update_task_status_from_assignments(db, task)
    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/tasks/{task.id}", status_code=303)


@app.get("/camps/{camp_id}/tasks/{task_id}/assignments/{assignment_id}/edit", response_class=HTMLResponse)
async def edit_task_assignment_form(
    request: Request,
    camp_id: int,
    task_id: int,
    assignment_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    task = (
        db.query(Task)
        .filter(Task.id == task_id, Task.camp_id == camp.id)
        .first()
    )

    if task is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Task not found."},
            status_code=404,
        )

    assignment = (
        db.query(TaskAssignment)
        .filter(
            TaskAssignment.id == assignment_id,
            TaskAssignment.task_id == task.id,
            TaskAssignment.camp_id == camp.id,
        )
        .first()
    )

    if assignment is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Task assignment not found."},
            status_code=404,
        )

    people = (
        db.query(Person)
        .filter(Person.camp_id == camp.id)
        .order_by(Person.person_type, Person.last_name, Person.first_name)
        .all()
    )

    teams = (
        db.query(Team)
        .filter(Team.camp_id == camp.id)
        .order_by(Team.team_type, Team.name)
        .all()
    )

    return templates.TemplateResponse(
        "tasks/edit_assignment.html",
        {
            "request": request,
            "camp": camp,
            "task": task,
            "assignment": assignment,
            "people": people,
            "teams": teams,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/tasks/{task_id}/assignments/{assignment_id}/edit")
async def update_task_assignment(
    request: Request,
    camp_id: int,
    task_id: int,
    assignment_id: int,
    assignee: str = Form(...),
    status_override: str = Form(""),
    assignment_notes: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    task = (
        db.query(Task)
        .filter(Task.id == task_id, Task.camp_id == camp.id)
        .first()
    )

    if task is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Task not found."},
            status_code=404,
        )

    assignment = (
        db.query(TaskAssignment)
        .filter(
            TaskAssignment.id == assignment_id,
            TaskAssignment.task_id == task.id,
            TaskAssignment.camp_id == camp.id,
        )
        .first()
    )

    if assignment is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Task assignment not found."},
            status_code=404,
        )

    kind, raw_id = assignee.split(":", 1)
    target_id = int(raw_id)

    new_assigned_person_id = None
    new_assigned_team_id = None

    if kind == "person":
        person = (
            db.query(Person)
            .filter(Person.id == target_id, Person.camp_id == camp.id)
            .first()
        )

        if person is None:
            return templates.TemplateResponse(
                "not_found.html",
                {"request": request, "message": "Person not found."},
                status_code=404,
            )

        new_assigned_person_id = person.id

    elif kind == "team":
        team = (
            db.query(Team)
            .filter(Team.id == target_id, Team.camp_id == camp.id)
            .first()
        )

        if team is None:
            return templates.TemplateResponse(
                "not_found.html",
                {"request": request, "message": "Team not found."},
                status_code=404,
            )

        new_assigned_team_id = team.id

    else:
        people = db.query(Person).filter(Person.camp_id == camp.id).all()
        teams = db.query(Team).filter(Team.camp_id == camp.id).all()
        return templates.TemplateResponse(
            "tasks/edit_assignment.html",
            {
                "request": request,
                "camp": camp,
                "task": task,
                "assignment": assignment,
                "people": people,
                "teams": teams,
                "error": "Choose a valid person or team.",
            },
            status_code=400,
        )

    if task_assignment_duplicate_exists(
        db,
        task,
        assigned_person_id=new_assigned_person_id,
        assigned_team_id=new_assigned_team_id,
        exclude_assignment_id=assignment.id,
    ):
        people = db.query(Person).filter(Person.camp_id == camp.id).order_by(Person.person_type, Person.last_name, Person.first_name).all()
        teams = db.query(Team).filter(Team.camp_id == camp.id).order_by(Team.team_type, Team.name).all()

        return templates.TemplateResponse(
            "tasks/edit_assignment.html",
            {
                "request": request,
                "camp": camp,
                "task": task,
                "assignment": assignment,
                "people": people,
                "teams": teams,
                "error": "This task is already assigned to that person or team.",
            },
            status_code=400,
        )

    assignment.assigned_person_id = new_assigned_person_id
    assignment.assigned_team_id = new_assigned_team_id
    assignment.status_override = status_override.strip() or None
    assignment.assignment_notes = assignment_notes.strip() or None

    update_task_status_from_assignments(db, task)
    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/tasks/{task.id}", status_code=303)


@app.post("/camps/{camp_id}/tasks/{task_id}/assignments/{assignment_id}/remove")
async def remove_task_assignment(
    camp_id: int,
    task_id: int,
    assignment_id: int,
    db: Session = Depends(get_db),
):
    task = (
        db.query(Task)
        .filter(Task.id == task_id, Task.camp_id == camp_id)
        .first()
    )

    assignment = (
        db.query(TaskAssignment)
        .filter(
            TaskAssignment.id == assignment_id,
            TaskAssignment.camp_id == camp_id,
            TaskAssignment.task_id == task_id,
        )
        .first()
    )

    if assignment is not None:
        db.delete(assignment)
        db.flush()

        if task is not None:
            update_task_status_from_assignments(db, task)

        db.commit()

    return RedirectResponse(url=f"/camps/{camp_id}/tasks/{task_id}", status_code=303)


@app.get("/readiness", response_class=HTMLResponse)
async def readiness_page(request: Request):
    return templates.TemplateResponse("readiness.html", {"request": request})


@app.get("/outputs", response_class=HTMLResponse)
async def outputs_page(request: Request):
    return templates.TemplateResponse("outputs.html", {"request": request})


@app.get("/health")
async def health():
    return {"status": "ok"}



def build_task_print_rows(db: Session, camp: Camp, tasks: list[Task], source_label: str = ""):
    rows = []

    for task in tasks:
        assignment_count = (
            db.query(TaskAssignment)
            .filter(TaskAssignment.camp_id == camp.id, TaskAssignment.task_id == task.id)
            .count()
        )

        rows.append(
            {
                "task": task,
                "source": source_label,
                "assignment_count": assignment_count,
            }
        )

    return rows


def order_tasks_for_print(query):
    return query.order_by(
        Task.phase,
        Task.category,
        Task.priority,
        Task.due_date,
        Task.title,
    )



def count_tasks_for_person(db: Session, camp: Camp, person: Person):
    direct_task_ids = {
        row.task_id
        for row in db.query(TaskAssignment)
        .filter(
            TaskAssignment.camp_id == camp.id,
            TaskAssignment.assigned_person_id == person.id,
        )
        .all()
    }

    team_ids = {
        row.team_id
        for row in db.query(TeamMembership)
        .filter(
            TeamMembership.camp_id == camp.id,
            TeamMembership.person_id == person.id,
        )
        .all()
    }

    team_task_ids = set()

    if team_ids:
        team_task_ids = {
            row.task_id
            for row in db.query(TaskAssignment)
            .filter(
                TaskAssignment.camp_id == camp.id,
                TaskAssignment.assigned_team_id.in_(team_ids),
            )
            .all()
        }

    return len(direct_task_ids | team_task_ids)


def count_tasks_for_team(db: Session, camp: Camp, team: Team):
    return (
        db.query(TaskAssignment)
        .filter(
            TaskAssignment.camp_id == camp.id,
            TaskAssignment.assigned_team_id == team.id,
        )
        .count()
    )


def count_tasks_for_phase(db: Session, camp: Camp, phase: TaskPhase):
    return (
        db.query(Task)
        .filter(Task.camp_id == camp.id, Task.phase == phase.name)
        .count()
    )


def count_tasks_for_category(db: Session, camp: Camp, category: TaskCategory):
    return (
        db.query(Task)
        .filter(Task.camp_id == camp.id, Task.category == category.name)
        .count()
    )


@app.get("/camps/{camp_id}/task-sheets", response_class=HTMLResponse)
async def task_sheet_hub(request: Request, camp_id: int, db: Session = Depends(get_db)):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    people = (
        db.query(Person)
        .filter(Person.camp_id == camp.id)
        .order_by(Person.person_type, Person.last_name, Person.first_name)
        .all()
    )

    teams = (
        db.query(Team)
        .filter(Team.camp_id == camp.id)
        .order_by(Team.team_type, Team.name)
        .all()
    )

    phases = get_active_task_phases(db, camp)
    categories = get_active_task_categories(db, camp)

    all_tasks = db.query(Task).filter(Task.camp_id == camp.id).all()
    assignment_counts = {}

    for task in all_tasks:
        assignment_counts[task.id] = (
            db.query(TaskAssignment)
            .filter(TaskAssignment.camp_id == camp.id, TaskAssignment.task_id == task.id)
            .count()
        )

    unassigned_count = sum(1 for task in all_tasks if assignment_counts.get(task.id, 0) == 0)

    person_task_counts = {}

    for person in people:
        direct_task_ids = {
            row.task_id
            for row in db.query(TaskAssignment)
            .filter(
                TaskAssignment.camp_id == camp.id,
                TaskAssignment.assigned_person_id == person.id,
            )
            .all()
        }

        team_ids = {
            row.team_id
            for row in db.query(TeamMembership)
            .filter(
                TeamMembership.camp_id == camp.id,
                TeamMembership.person_id == person.id,
            )
            .all()
        }

        team_task_ids = set()

        if team_ids:
            team_task_ids = {
                row.task_id
                for row in db.query(TaskAssignment)
                .filter(
                    TaskAssignment.camp_id == camp.id,
                    TaskAssignment.assigned_team_id.in_(team_ids),
                )
                .all()
            }

        person_task_counts[person.id] = len(direct_task_ids | team_task_ids)

    team_task_counts = {}

    for team in teams:
        team_task_counts[team.id] = (
            db.query(TaskAssignment)
            .filter(
                TaskAssignment.camp_id == camp.id,
                TaskAssignment.assigned_team_id == team.id,
            )
            .count()
        )

    phase_task_counts = {}

    for phase in phases:
        phase_task_counts[phase.id] = (
            db.query(Task)
            .filter(Task.camp_id == camp.id, Task.phase == phase.name)
            .count()
        )

    category_task_counts = {
        category.id: count_tasks_for_category(db, camp, category)
        for category in categories
    }

    return templates.TemplateResponse(
        "task_sheets/hub.html",
        {
            "request": request,
            "camp": camp,
            "people": people,
            "teams": teams,
            "phases": phases,
            "categories": categories,
            "unassigned_count": unassigned_count,
            "person_task_counts": person_task_counts,
            "team_task_counts": team_task_counts,
            "phase_task_counts": phase_task_counts,
            "category_task_counts": category_task_counts,
        },
    )


@app.get("/camps/{camp_id}/task-sheets/unassigned", response_class=HTMLResponse)
async def unassigned_task_sheet(request: Request, camp_id: int, db: Session = Depends(get_db)):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    all_tasks = order_tasks_for_print(
        db.query(Task).filter(Task.camp_id == camp.id)
    ).all()

    unassigned_tasks = []

    for task in all_tasks:
        assignment_count = (
            db.query(TaskAssignment)
            .filter(TaskAssignment.camp_id == camp.id, TaskAssignment.task_id == task.id)
            .count()
        )

        if assignment_count == 0:
            unassigned_tasks.append(task)

    rows = build_task_print_rows(db, camp, unassigned_tasks, "Needs owner")

    return templates.TemplateResponse(
        "task_sheets/print.html",
        {
            "request": request,
            "camp": camp,
            "title": "Unassigned Task Sheet",
            "subtitle": "Tasks that still need an owner.",
            "rows": rows,
            "back_url": f"/camps/{camp.id}/task-sheets",
            "empty_message": "No unassigned tasks found.",
        },
    )


@app.get("/camps/{camp_id}/task-sheets/person/{person_id}", response_class=HTMLResponse)
async def person_task_sheet(
    request: Request,
    camp_id: int,
    person_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    person = (
        db.query(Person)
        .filter(Person.id == person_id, Person.camp_id == camp.id)
        .first()
    )

    if person is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Person not found."},
            status_code=404,
        )

    direct_rows = (
        db.query(TaskAssignment, Task)
        .join(Task, Task.id == TaskAssignment.task_id)
        .filter(
            TaskAssignment.camp_id == camp.id,
            TaskAssignment.assigned_person_id == person.id,
        )
        .order_by(Task.phase, Task.category, Task.priority, Task.due_date, Task.title)
        .all()
    )

    team_memberships = (
        db.query(TeamMembership, Team)
        .join(Team, Team.id == TeamMembership.team_id)
        .filter(
            TeamMembership.camp_id == camp.id,
            TeamMembership.person_id == person.id,
        )
        .order_by(Team.name)
        .all()
    )

    team_ids = [team.id for membership, team in team_memberships]

    team_rows = []

    if team_ids:
        team_rows = (
            db.query(TaskAssignment, Task, Team)
            .join(Task, Task.id == TaskAssignment.task_id)
            .join(Team, Team.id == TaskAssignment.assigned_team_id)
            .filter(
                TaskAssignment.camp_id == camp.id,
                TaskAssignment.assigned_team_id.in_(team_ids),
            )
            .order_by(Team.name, Task.phase, Task.category, Task.priority, Task.due_date, Task.title)
            .all()
        )

    rows = []

    seen_task_sources = set()

    for assignment, task in direct_rows:
        key = (task.id, "Direct")
        if key not in seen_task_sources:
            rows.extend(build_task_print_rows(db, camp, [task], "Direct assignment"))
            seen_task_sources.add(key)

    for assignment, task, team in team_rows:
        key = (task.id, f"Team:{team.id}")
        if key not in seen_task_sources:
            rows.extend(build_task_print_rows(db, camp, [task], f"Team: {team.name}"))
            seen_task_sources.add(key)

    return templates.TemplateResponse(
        "task_sheets/print.html",
        {
            "request": request,
            "camp": camp,
            "title": f"Task Sheet — {person.first_name} {person.last_name}",
            "subtitle": f"{person.person_type}. Includes direct tasks and tasks from their teams.",
            "rows": rows,
            "back_url": f"/camps/{camp.id}/task-sheets",
            "empty_message": "No tasks found for this person.",
        },
    )


@app.get("/camps/{camp_id}/task-sheets/team/{team_id}", response_class=HTMLResponse)
async def team_task_sheet(
    request: Request,
    camp_id: int,
    team_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    team = (
        db.query(Team)
        .filter(Team.id == team_id, Team.camp_id == camp.id)
        .first()
    )

    if team is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Team not found."},
            status_code=404,
        )

    task_rows = (
        db.query(TaskAssignment, Task)
        .join(Task, Task.id == TaskAssignment.task_id)
        .filter(
            TaskAssignment.camp_id == camp.id,
            TaskAssignment.assigned_team_id == team.id,
        )
        .order_by(Task.phase, Task.category, Task.priority, Task.due_date, Task.title)
        .all()
    )

    rows = []

    for assignment, task in task_rows:
        rows.extend(build_task_print_rows(db, camp, [task], f"Team: {team.name}"))

    members = (
        db.query(TeamMembership, Person)
        .join(Person, Person.id == TeamMembership.person_id)
        .filter(
            TeamMembership.camp_id == camp.id,
            TeamMembership.team_id == team.id,
        )
        .order_by(Person.person_type, Person.last_name, Person.first_name)
        .all()
    )

    return templates.TemplateResponse(
        "task_sheets/print.html",
        {
            "request": request,
            "camp": camp,
            "title": f"Team Task Sheet — {team.name}",
            "subtitle": f"{team.team_type}. Includes tasks assigned to this team.",
            "rows": rows,
            "members": members,
            "back_url": f"/camps/{camp.id}/task-sheets",
            "empty_message": "No tasks found for this team.",
        },
    )


@app.get("/camps/{camp_id}/task-sheets/phase/{phase_id}", response_class=HTMLResponse)
async def phase_task_sheet(
    request: Request,
    camp_id: int,
    phase_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    phase = (
        db.query(TaskPhase)
        .filter(TaskPhase.id == phase_id, TaskPhase.camp_id == camp.id)
        .first()
    )

    if phase is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Task phase not found."},
            status_code=404,
        )

    tasks = order_tasks_for_print(
        db.query(Task).filter(Task.camp_id == camp.id, Task.phase == phase.name)
    ).all()

    rows = build_task_print_rows(db, camp, tasks, f"Phase: {phase.name}")

    return templates.TemplateResponse(
        "task_sheets/print.html",
        {
            "request": request,
            "camp": camp,
            "title": f"Phase Task Sheet — {phase.name}",
            "subtitle": phase.description or "Tasks grouped by planning or operational phase.",
            "rows": rows,
            "back_url": f"/camps/{camp.id}/task-sheets",
            "empty_message": "No tasks found for this phase.",
        },
    )



@app.get("/camps/{camp_id}/task-sheets/category/{category_id}", response_class=HTMLResponse)
async def category_task_sheet(
    request: Request,
    camp_id: int,
    category_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    category = (
        db.query(TaskCategory)
        .filter(TaskCategory.id == category_id, TaskCategory.camp_id == camp.id)
        .first()
    )

    if category is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Task category not found."},
            status_code=404,
        )

    tasks = order_tasks_for_print(
        db.query(Task).filter(Task.camp_id == camp.id, Task.category == category.name)
    ).all()

    rows = build_task_print_rows(db, camp, tasks, f"Category: {category.name}")

    return templates.TemplateResponse(
        "task_sheets/print.html",
        {
            "request": request,
            "camp": camp,
            "title": f"Category Task Sheet — {category.name}",
            "subtitle": category.description or "Tasks grouped by work area.",
            "rows": rows,
            "back_url": f"/camps/{camp.id}/task-sheets",
            "empty_message": "No tasks found for this category.",
        },
    )




@app.get("/camps/{camp_id}/activities", response_class=HTMLResponse)
async def camp_activity_list(request: Request, camp_id: int, db: Session = Depends(get_db)):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    activities = (
        db.query(Activity)
        .filter(Activity.camp_id == camp.id)
        .order_by(Activity.name)
        .all()
    )

    people = get_people_for_activity_leads(db, camp)
    lead_names = {person.id: person_display_name(person) for person in people}

    risk_statuses = {}

    for activity in activities:
        risk_assessment = (
            db.query(ActivityRiskAssessment)
            .filter(
                ActivityRiskAssessment.camp_id == camp.id,
                ActivityRiskAssessment.activity_id == activity.id,
            )
            .first()
        )

        risk_statuses[activity.id] = risk_assessment.status if risk_assessment else "Not Started"

    return templates.TemplateResponse(
        "activities/list.html",
        {
            "request": request,
            "camp": camp,
            "activities": activities,
            "lead_names": lead_names,
            "risk_statuses": risk_statuses,
        },
    )


@app.get("/camps/{camp_id}/activities/new", response_class=HTMLResponse)
async def new_activity_form(request: Request, camp_id: int, db: Session = Depends(get_db)):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    people = get_people_for_activity_leads(db, camp)

    return templates.TemplateResponse(
        "activities/new.html",
        {
            "request": request,
            "camp": camp,
            "people": people,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/activities/new")
async def create_activity(
    request: Request,
    camp_id: int,
    name: str = Form(...),
    description: str = Form(""),
    default_duration_minutes: int = Form(60),
    default_location: str = Form(""),
    activity_lead_id: int = Form(0),
    supporting_adults_notes: str = Form(""),
    equipment_notes: str = Form(""),
    risk_notes: str = Form(""),
    wet_weather_alternative: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    people = get_people_for_activity_leads(db, camp)

    if not name.strip():
        return templates.TemplateResponse(
            "activities/new.html",
            {
                "request": request,
                "camp": camp,
                "people": people,
                "error": "Activity name is required.",
            },
            status_code=400,
        )

    lead_id = activity_lead_id or None

    if lead_id is not None:
        lead = (
            db.query(Person)
            .filter(Person.id == lead_id, Person.camp_id == camp.id)
            .first()
        )

        if lead is None:
            return templates.TemplateResponse(
                "activities/new.html",
                {
                    "request": request,
                    "camp": camp,
                    "people": people,
                    "error": "Selected activity lead does not belong to this camp.",
                },
                status_code=400,
            )

    activity = Activity(
        camp_id=camp.id,
        name=name.strip(),
        description=description.strip() or None,
        default_duration_minutes=default_duration_minutes,
        default_location=default_location.strip() or None,
        activity_lead_id=lead_id,
        supporting_adults_notes=supporting_adults_notes.strip() or None,
        equipment_notes=equipment_notes.strip() or None,
        risk_notes=risk_notes.strip() or None,
        wet_weather_alternative=wet_weather_alternative.strip() or None,
    )

    db.add(activity)
    db.commit()
    db.refresh(activity)

    return RedirectResponse(url=f"/camps/{camp.id}/activities/{activity.id}", status_code=303)


@app.get("/camps/{camp_id}/activities/{activity_id}/edit", response_class=HTMLResponse)
async def edit_activity_form(
    request: Request,
    camp_id: int,
    activity_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    activity = (
        db.query(Activity)
        .filter(Activity.id == activity_id, Activity.camp_id == camp.id)
        .first()
    )

    if activity is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Activity not found."},
            status_code=404,
        )

    people = get_people_for_activity_leads(db, camp)

    return templates.TemplateResponse(
        "activities/edit.html",
        {
            "request": request,
            "camp": camp,
            "activity": activity,
            "people": people,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/activities/{activity_id}/edit")
async def update_activity(
    request: Request,
    camp_id: int,
    activity_id: int,
    name: str = Form(...),
    description: str = Form(""),
    default_duration_minutes: int = Form(60),
    default_location: str = Form(""),
    activity_lead_id: int = Form(0),
    supporting_adults_notes: str = Form(""),
    equipment_notes: str = Form(""),
    risk_notes: str = Form(""),
    wet_weather_alternative: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    activity = (
        db.query(Activity)
        .filter(Activity.id == activity_id, Activity.camp_id == camp.id)
        .first()
    )

    if activity is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Activity not found."},
            status_code=404,
        )

    people = get_people_for_activity_leads(db, camp)

    if not name.strip():
        return templates.TemplateResponse(
            "activities/edit.html",
            {
                "request": request,
                "camp": camp,
                "activity": activity,
                "people": people,
                "error": "Activity name is required.",
            },
            status_code=400,
        )

    lead_id = activity_lead_id or None

    if lead_id is not None:
        lead = (
            db.query(Person)
            .filter(Person.id == lead_id, Person.camp_id == camp.id)
            .first()
        )

        if lead is None:
            return templates.TemplateResponse(
                "activities/edit.html",
                {
                    "request": request,
                    "camp": camp,
                    "activity": activity,
                    "people": people,
                    "error": "Selected activity lead does not belong to this camp.",
                },
                status_code=400,
            )

    activity.name = name.strip()
    activity.description = description.strip() or None
    activity.default_duration_minutes = default_duration_minutes
    activity.default_location = default_location.strip() or None
    activity.activity_lead_id = lead_id
    activity.supporting_adults_notes = supporting_adults_notes.strip() or None
    activity.equipment_notes = equipment_notes.strip() or None
    activity.risk_notes = risk_notes.strip() or None
    activity.wet_weather_alternative = wet_weather_alternative.strip() or None

    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/activities/{activity.id}", status_code=303)


@app.post("/camps/{camp_id}/activities/{activity_id}/delete")
async def delete_activity(
    request: Request,
    camp_id: int,
    activity_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    activity = (
        db.query(Activity)
        .filter(Activity.id == activity_id, Activity.camp_id == camp.id)
        .first()
    )

    if activity is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Activity not found."},
            status_code=404,
        )

    db.delete(activity)
    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/activities", status_code=303)


@app.get("/camps/{camp_id}/activities/{activity_id}", response_class=HTMLResponse)
async def activity_detail(
    request: Request,
    camp_id: int,
    activity_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    activity = (
        db.query(Activity)
        .filter(Activity.id == activity_id, Activity.camp_id == camp.id)
        .first()
    )

    if activity is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Activity not found."},
            status_code=404,
        )

    lead = None

    if activity.activity_lead_id:
        lead = (
            db.query(Person)
            .filter(Person.id == activity.activity_lead_id, Person.camp_id == camp.id)
            .first()
        )

    risk_assessment = (
        db.query(ActivityRiskAssessment)
        .filter(
            ActivityRiskAssessment.camp_id == camp.id,
            ActivityRiskAssessment.activity_id == activity.id,
        )
        .first()
    )

    risk_status = risk_assessment.status if risk_assessment else "Not Started"

    return templates.TemplateResponse(
        "activities/detail.html",
        {
            "request": request,
            "camp": camp,
            "activity": activity,
            "lead": lead,
            "risk_status": risk_status,
        },
    )



@app.get("/camps/{camp_id}/risk-assessments", response_class=HTMLResponse)
async def risk_assessment_hub(request: Request, camp_id: int, db: Session = Depends(get_db)):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    camp_risk = ensure_camp_risk_assessment(db, camp)

    activities = (
        db.query(Activity)
        .filter(Activity.camp_id == camp.id)
        .order_by(Activity.name)
        .all()
    )

    activity_risks = {}

    for activity in activities:
        activity_risks[activity.id] = ensure_activity_risk_assessment(db, camp, activity)

    return templates.TemplateResponse(
        "risk_assessments/hub.html",
        {
            "request": request,
            "camp": camp,
            "camp_risk": camp_risk,
            "activities": activities,
            "activity_risks": activity_risks,
        },
    )


@app.get("/camps/{camp_id}/risk-assessments/camp", response_class=HTMLResponse)
async def camp_risk_assessment_detail(
    request: Request,
    camp_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    risk_assessment = ensure_camp_risk_assessment(db, camp)

    controls = (
        db.query(CampRiskControl)
        .filter(CampRiskControl.risk_assessment_id == risk_assessment.id)
        .order_by(CampRiskControl.sort_order, CampRiskControl.id)
        .all()
    )

    return templates.TemplateResponse(
        "risk_assessments/camp.html",
        {
            "request": request,
            "camp": camp,
            "risk_assessment": risk_assessment,
            "controls": controls,
            "statuses": RISK_STATUSES,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/risk-assessments/camp")
async def update_camp_risk_assessment(
    request: Request,
    camp_id: int,
    title: str = Form(...),
    status: str = Form("Not Started"),
    prepared_by: str = Form(""),
    review_date: str = Form(""),
    submitted_date: str = Form(""),
    approved_date: str = Form(""),
    approval_notes: str = Form(""),
    site_notes: str = Form(""),
    overnight_notes: str = Form(""),
    supervision_notes: str = Form(""),
    emergency_notes: str = Form(""),
    communication_notes: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    risk_assessment = ensure_camp_risk_assessment(db, camp)

    risk_assessment.title = title.strip() or f"{camp.name} Risk Assessment"
    risk_assessment.status = status
    risk_assessment.prepared_by = prepared_by.strip() or None
    risk_assessment.review_date = parse_optional_date(review_date)
    risk_assessment.submitted_date = parse_optional_date(submitted_date)
    risk_assessment.approved_date = parse_optional_date(approved_date)
    risk_assessment.approval_notes = approval_notes.strip() or None
    risk_assessment.site_notes = site_notes.strip() or None
    risk_assessment.overnight_notes = overnight_notes.strip() or None
    risk_assessment.supervision_notes = supervision_notes.strip() or None
    risk_assessment.emergency_notes = emergency_notes.strip() or None
    risk_assessment.communication_notes = communication_notes.strip() or None

    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/risk-assessments/camp", status_code=303)


@app.post("/camps/{camp_id}/risk-assessments/camp/controls/new")
async def add_camp_risk_control(
    request: Request,
    camp_id: int,
    hazard: str = Form(...),
    who_might_be_harmed: str = Form(""),
    controls_in_place: str = Form(""),
    further_controls_needed: str = Form(""),
    responsible_person: str = Form(""),
    review_notes: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    risk_assessment = ensure_camp_risk_assessment(db, camp)

    if hazard.strip():
        next_order = (
            db.query(CampRiskControl)
            .filter(CampRiskControl.risk_assessment_id == risk_assessment.id)
            .count()
            + 1
        )

        db.add(
            CampRiskControl(
                risk_assessment_id=risk_assessment.id,
                sort_order=next_order,
                hazard=hazard.strip(),
                who_might_be_harmed=who_might_be_harmed.strip() or None,
                controls_in_place=controls_in_place.strip() or None,
                further_controls_needed=further_controls_needed.strip() or None,
                responsible_person=responsible_person.strip() or None,
                review_notes=review_notes.strip() or None,
            )
        )

        db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/risk-assessments/camp", status_code=303)


@app.post("/camps/{camp_id}/risk-assessments/camp/controls/{control_id}/delete")
async def delete_camp_risk_control(
    request: Request,
    camp_id: int,
    control_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    risk_assessment = ensure_camp_risk_assessment(db, camp)

    control = (
        db.query(CampRiskControl)
        .filter(
            CampRiskControl.id == control_id,
            CampRiskControl.risk_assessment_id == risk_assessment.id,
        )
        .first()
    )

    if control:
        db.delete(control)
        db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/risk-assessments/camp", status_code=303)


@app.get("/camps/{camp_id}/activities/{activity_id}/risk-assessment", response_class=HTMLResponse)
async def activity_risk_assessment_detail(
    request: Request,
    camp_id: int,
    activity_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    activity = (
        db.query(Activity)
        .filter(Activity.id == activity_id, Activity.camp_id == camp.id)
        .first()
    )

    if activity is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Activity not found."},
            status_code=404,
        )

    risk_assessment = ensure_activity_risk_assessment(db, camp, activity)

    controls = (
        db.query(ActivityRiskControl)
        .filter(ActivityRiskControl.risk_assessment_id == risk_assessment.id)
        .order_by(ActivityRiskControl.sort_order, ActivityRiskControl.id)
        .all()
    )

    return templates.TemplateResponse(
        "risk_assessments/activity.html",
        {
            "request": request,
            "camp": camp,
            "activity": activity,
            "risk_assessment": risk_assessment,
            "controls": controls,
            "statuses": RISK_STATUSES,
            "source_types": ACTIVITY_RISK_SOURCE_TYPES,
            "error": None,
        },
    )


@app.post("/camps/{camp_id}/activities/{activity_id}/risk-assessment")
async def update_activity_risk_assessment(
    request: Request,
    camp_id: int,
    activity_id: int,
    source_type: str = Form("Created in app"),
    status: str = Form("Not Started"),
    leader_in_charge: str = Form(""),
    prepared_by: str = Form(""),
    review_date: str = Form(""),
    provider_name: str = Form(""),
    provider_contact: str = Form(""),
    provider_reference: str = Form(""),
    provider_risk_assessment_received: str = Form("off"),
    provider_insurance_checked: str = Form("off"),
    provider_qualification_checked: str = Form("off"),
    scout_led_parts_notes: str = Form(""),
    overall_notes: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    activity = (
        db.query(Activity)
        .filter(Activity.id == activity_id, Activity.camp_id == camp.id)
        .first()
    )

    if activity is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Activity not found."},
            status_code=404,
        )

    risk_assessment = ensure_activity_risk_assessment(db, camp, activity)

    risk_assessment.source_type = source_type
    risk_assessment.status = status
    risk_assessment.leader_in_charge = leader_in_charge.strip() or None
    risk_assessment.prepared_by = prepared_by.strip() or None
    risk_assessment.review_date = parse_optional_date(review_date)
    risk_assessment.provider_name = provider_name.strip() or None
    risk_assessment.provider_contact = provider_contact.strip() or None
    risk_assessment.provider_reference = provider_reference.strip() or None
    risk_assessment.provider_risk_assessment_received = provider_risk_assessment_received == "on"
    risk_assessment.provider_insurance_checked = provider_insurance_checked == "on"
    risk_assessment.provider_qualification_checked = provider_qualification_checked == "on"
    risk_assessment.scout_led_parts_notes = scout_led_parts_notes.strip() or None
    risk_assessment.overall_notes = overall_notes.strip() or None

    db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/activities/{activity.id}/risk-assessment", status_code=303)


@app.post("/camps/{camp_id}/activities/{activity_id}/risk-assessment/controls/new")
async def add_activity_risk_control(
    request: Request,
    camp_id: int,
    activity_id: int,
    hazard: str = Form(...),
    who_might_be_harmed: str = Form(""),
    controls_in_place: str = Form(""),
    further_controls_needed: str = Form(""),
    responsible_person: str = Form(""),
    review_notes: str = Form(""),
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    activity = (
        db.query(Activity)
        .filter(Activity.id == activity_id, Activity.camp_id == camp.id)
        .first()
    )

    if activity is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Activity not found."},
            status_code=404,
        )

    risk_assessment = ensure_activity_risk_assessment(db, camp, activity)

    if hazard.strip():
        next_order = (
            db.query(ActivityRiskControl)
            .filter(ActivityRiskControl.risk_assessment_id == risk_assessment.id)
            .count()
            + 1
        )

        db.add(
            ActivityRiskControl(
                risk_assessment_id=risk_assessment.id,
                sort_order=next_order,
                hazard=hazard.strip(),
                who_might_be_harmed=who_might_be_harmed.strip() or None,
                controls_in_place=controls_in_place.strip() or None,
                further_controls_needed=further_controls_needed.strip() or None,
                responsible_person=responsible_person.strip() or None,
                review_notes=review_notes.strip() or None,
            )
        )

        db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/activities/{activity.id}/risk-assessment", status_code=303)


@app.post("/camps/{camp_id}/activities/{activity_id}/risk-assessment/controls/{control_id}/delete")
async def delete_activity_risk_control(
    request: Request,
    camp_id: int,
    activity_id: int,
    control_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    activity = (
        db.query(Activity)
        .filter(Activity.id == activity_id, Activity.camp_id == camp.id)
        .first()
    )

    if activity is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Activity not found."},
            status_code=404,
        )

    risk_assessment = ensure_activity_risk_assessment(db, camp, activity)

    control = (
        db.query(ActivityRiskControl)
        .filter(
            ActivityRiskControl.id == control_id,
            ActivityRiskControl.risk_assessment_id == risk_assessment.id,
        )
        .first()
    )

    if control:
        db.delete(control)
        db.commit()

    return RedirectResponse(url=f"/camps/{camp.id}/activities/{activity.id}/risk-assessment", status_code=303)



@app.get("/camps/{camp_id}/risk-assessments/camp/print", response_class=HTMLResponse)
async def print_camp_risk_assessment(
    request: Request,
    camp_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    risk_assessment = ensure_camp_risk_assessment(db, camp)

    controls = (
        db.query(CampRiskControl)
        .filter(CampRiskControl.risk_assessment_id == risk_assessment.id)
        .order_by(CampRiskControl.sort_order, CampRiskControl.id)
        .all()
    )

    return templates.TemplateResponse(
        "risk_assessments/camp_print.html",
        {
            "request": request,
            "camp": camp,
            "risk_assessment": risk_assessment,
            "controls": controls,
        },
    )


@app.get("/camps/{camp_id}/activities/{activity_id}/risk-assessment/print", response_class=HTMLResponse)
async def print_activity_risk_assessment(
    request: Request,
    camp_id: int,
    activity_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    activity = (
        db.query(Activity)
        .filter(Activity.id == activity_id, Activity.camp_id == camp.id)
        .first()
    )

    if activity is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Activity not found."},
            status_code=404,
        )

    risk_assessment = ensure_activity_risk_assessment(db, camp, activity)

    controls = (
        db.query(ActivityRiskControl)
        .filter(ActivityRiskControl.risk_assessment_id == risk_assessment.id)
        .order_by(ActivityRiskControl.sort_order, ActivityRiskControl.id)
        .all()
    )

    lead = None

    if activity.activity_lead_id:
        lead = (
            db.query(Person)
            .filter(Person.id == activity.activity_lead_id, Person.camp_id == camp.id)
            .first()
        )

    return templates.TemplateResponse(
        "risk_assessments/activity_print.html",
        {
            "request": request,
            "camp": camp,
            "activity": activity,
            "lead": lead,
            "risk_assessment": risk_assessment,
            "controls": controls,
        },
    )



@app.get("/camps/{camp_id}/risk-assessments/print-pack", response_class=HTMLResponse)
async def print_risk_assessment_pack(
    request: Request,
    camp_id: int,
    db: Session = Depends(get_db),
):
    camp = db.get(Camp, camp_id)

    if camp is None:
        return templates.TemplateResponse(
            "not_found.html",
            {"request": request, "message": "Camp not found."},
            status_code=404,
        )

    camp_risk = ensure_camp_risk_assessment(db, camp)

    camp_controls = (
        db.query(CampRiskControl)
        .filter(CampRiskControl.risk_assessment_id == camp_risk.id)
        .order_by(CampRiskControl.sort_order, CampRiskControl.id)
        .all()
    )

    activities = (
        db.query(Activity)
        .filter(Activity.camp_id == camp.id)
        .order_by(Activity.name)
        .all()
    )

    activity_sections = []
    provider_sections = []

    for activity in activities:
        risk_assessment = ensure_activity_risk_assessment(db, camp, activity)

        controls = (
            db.query(ActivityRiskControl)
            .filter(ActivityRiskControl.risk_assessment_id == risk_assessment.id)
            .order_by(ActivityRiskControl.sort_order, ActivityRiskControl.id)
            .all()
        )

        lead = None

        if activity.activity_lead_id:
            lead = (
                db.query(Person)
                .filter(Person.id == activity.activity_lead_id, Person.camp_id == camp.id)
                .first()
            )

        section = {
            "activity": activity,
            "risk_assessment": risk_assessment,
            "controls": controls,
            "lead": lead,
        }

        activity_sections.append(section)

        if risk_assessment.source_type == "External provider":
            provider_sections.append(section)

    status_counts = {
        "Not Started": 0,
        "Draft": 0,
        "Ready for Review": 0,
        "Submitted": 0,
        "Approved": 0,
        "Needs Update": 0,
    }

    status_counts[camp_risk.status] = status_counts.get(camp_risk.status, 0) + 1

    for section in activity_sections:
        status = section["risk_assessment"].status
        status_counts[status] = status_counts.get(status, 0) + 1

    return templates.TemplateResponse(
        "risk_assessments/pack_print.html",
        {
            "request": request,
            "camp": camp,
            "camp_risk": camp_risk,
            "camp_controls": camp_controls,
            "activity_sections": activity_sections,
            "provider_sections": provider_sections,
            "status_counts": status_counts,
        },
    )

